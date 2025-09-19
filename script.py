import re
import time
import io
import streamlit as st
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook 
from openpyxl.utils import get_column_letter # Add this import
from urllib.parse import quote_plus

TOKEN_RE = re.compile(r"""
    \d+(?:\.\d+)?            # integer or decimal, e.g. 3 or 3.5
    (?:\s*(?:g|mg|oz))?      # optional unit, allows a space: "3.5g" or "3.5 g"
  | [A-Za-z]+                # or plain words
""", re.VERBOSE | re.IGNORECASE)

# common words to ignore
STOPWORDS = {
    'a','an','and','at','by','for','in','of','on','or','the','to','with', 'sample', 'hybrid', 'indica', 'sativa', 'pre', 'pod', 'popcorn', 'shake', 'pills'
}

# Define a list of common flavors
FLAVOR_LIST = [
    'apple', 'banana', 'berry', 'raspberry', 'blueberry', 'bubblegum', 'cherry', 'chocolate',
    'citrus', 'cinnamon', 'coffee', 'cookies', 'cream', 'diesel', 'fruit', 'grape', 'lemon',
    'lime', 'mango', 'mint', 'orange', 'peach', 'pineapple', 'lemonade', 'sour', 'strawberry', 
    'tropical', 'vanilla', 'watermelon', 'zesty', 'sweet', 'peppermint', 'spearmint', 
    'grapefruit', 'guava', 'spicy', 'woody', 'floral', 'gelato', 'gsc', 'haze', 
    'zkittlez', 'runtz', 'mac', 'purp', 'gg4', 'gmo', 'shake', 'popcorn', 'mimosa'
]

# Category Mapping between Excel Categories and Website Categories
category_mapping = {
    'BEVERAGE': 'Edibles',         # BEVERAGE to Edible on the website
    'EDIBLE': 'Edibles',           # EDIBLE to Edible on the website
    'PILL': 'Edibles',             # PILL to Edible on the website
    'FLOWER': 'Flower',            # FLOWER to Flower on the website
    'CARTRIDGE': 'Vaporizers',     # CARTRIDGE to Vaporizers on the website
    'EXTRACT': 'Concentrates',     # EXTRACT to Concentrates on the website
    'TOPICAL': 'Topicals',         # TOPICAL to Topicals on the website
    'PREROLL': 'Pre-Rolls',        # PREROLL to Pre-Rolls on the website
    'TINCTURE': 'Tinctures',       # TINCTURE to Tinctures on the website
    'CBD': 'Tinctures',            # CBD to Tinctures on the website
    'MERCH': 'Apparel'             # MERCH to Apparel on the website
    # Add more mappings as needed...
}

# Categories on the site that have *no* weight filter
no_weight_categories = ['Edibles', 'Topicals', 'Accessories', 'Apparel']

# Global buffer to hold the Excel file in memory
excel_buffer = None

def save_updated_excel_to_memory(uploaded_file):
    """
    Loads the uploaded Excel file into an in-memory BytesIO buffer.
    This buffer will be used and updated throughout the scraping process.
    """
    global excel_buffer
    excel_buffer = io.BytesIO(uploaded_file.getvalue())
    print("Excel file loaded into memory for updates.")

def save_data_to_file(row_index, discounted_price, original_price, product_thc, product_url):
    """
    Updates a specific row in the in-memory Excel workbook with scraped data.
    """
    global excel_buffer
    if excel_buffer is None:
        st.error("Error: Excel buffer not initialized. Please upload a file first.")
        return

    try:
        # Load the workbook from the in-memory BytesIO object
        excel_buffer.seek(0) # Go to the beginning of the buffer
        wb = load_workbook(excel_buffer)
        
        # Select the specific sheet named "Pricing Research"
        sheet_name = "Pricing Research"
        if sheet_name not in wb.sheetnames:
            ui_log(f"Warning: Sheet '{sheet_name}' not found. Creating it.")
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # openpyxl uses 1-based indexing for rows and columns
        # The row_index from pandas is 0-based, so add 2 for Excel (1 for header, 1 for 0-base to 1-base)
        excel_row = row_index + 2

        # Define the target columns (AS, AT, AU, AV)
        # AY is column 51, AZ is 52, BA is 53, BB is 54
        col_original_price = get_column_letter(51)   # AY
        col_discounted_price = get_column_letter(52) # AZ
        col_thc_content = get_column_letter(53)      # BA
        col_product_url = get_column_letter(54)      # BB

        # Handle multiple values by joining them with commas
        ws[f"{col_discounted_price}{excel_row}"] = ", ".join(map(str, discounted_price)) if isinstance(discounted_price, list) else discounted_price
        ws[f"{col_original_price}{excel_row}"] = ", ".join(map(str, original_price)) if isinstance(original_price, list) else original_price
        ws[f"{col_thc_content}{excel_row}"] = ", ".join(map(str, product_thc)) if isinstance(product_thc, list) else product_thc
        ws[f"{col_product_url}{excel_row}"] = ", ".join(map(str, product_url)) if isinstance(product_url, list) else product_url

        # Save the modified workbook back to the BytesIO buffer
        new_buffer = io.BytesIO()
        wb.save(new_buffer)

        # Reset pointer so future reads start from the beginning
        new_buffer.seek(0)

        # Update the global buffer
        excel_buffer = new_buffer

        print(f"Row {excel_row} updated in memory for product at index {row_index}.")

    except Exception as e:
        st.error(f"Error saving data to Excel for row {row_index}: {e}")



from playwright.sync_api import sync_playwright

# from playwright.sync_api import sync_playwright

def get_driver():
    """
    Initialize and configure the Playwright WebDriver.
    """
    with sync_playwright() as p:
        # Launch the Chromium browser
        browser = p.chromium.launch(headless=True)  # or p.firefox.launch(headless=True)
        page = browser.new_page()
        
        # Set the page as the return value
        return page



def clean_thc_value(thc_string):
    """
    Extracts only the value part from a THC string like 'THC: 100 mg'.
    Returns 'N/A' if the input is 'N/A' or doesn't match the expected format.
    """
    if thc_string == "N/A":
        return "N/A"

    match = re.search(r"THC:\s*(.*)", thc_string, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return thc_string # Return original if no match (e.g., if it's already just the value)

def handle_age_verification_popup(driver, wait):
    """
    Handles the age verification pop-up on the Terrabis Grayville Dispensary website.
    Clicks the "yes, I'm 21 or older" button if the pop-up appears.
    """
    ui_log("Waiting for pop-up to appear and attempting to close it...")
    time.sleep(4)  # Initial wait to allow the page to load
    try:
        # Wait for the "yes, I'm 21 or older" button to be clickable
        age_verification_button = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.pum-close.elementor-element-ebd2f15"))
        )
        age_verification_button.click()  # Click the age verification button
        print("Pop-up closed successfully!")
        time.sleep(3)  # Increased delay after closing pop-up to allow page to settle
    except Exception as popup_e:
        print(f"Could not find or click the pop-up close button. It might not have appeared or its locator changed. Error: {popup_e}")
        ui_log("Continuing without closing pop-up. You might need to close it manually.")

# Define direct links for specific categories
direct_category_links = {
    'Merch': 'https://terrabis.co/order-online/grayville/?dtche%5Bsortby%5D=relevance&dtche%5Bcategory%5D=apparel',
    'Accessories': 'https://terrabis.co/order-online/grayville/?dtche%5Bsortby%5D=relevance&dtche%5Bcategory%5D=accessories'
}

def scrape_category(category, driver):
    website_url = 'https://terrabis.co/illinois/grayville/'
    
    # Check if a direct link is available for the current category
    # Use .capitalize() to match how categories are usually handled or mapped
    normalized_category = category.capitalize() 

    if normalized_category in direct_category_links:
        direct_link = direct_category_links[normalized_category]
        ui_log(f"Directly navigating to URL for category '{category}': {direct_link}")
        print(f"Directly navigating to URL for category '{category}': {direct_link}")
        driver.get(direct_link)
        wait = WebDriverWait(driver, 20)
        handle_age_verification_popup(driver, wait)
        print(f"Successfully navigated to direct link for category '{category}'.")
        ui_log(f"Category '{category}' selected via direct link.")
        time.sleep(3) # Give some time for the new category page to load
        return True

    # If no direct link, proceed with carousel navigation logic
    driver.get(website_url)

    wait = WebDriverWait(driver, 20) # Increased wait time for initial load
    handle_age_verification_popup(driver, wait)

    ui_log(f"Attempting to select category '{category}' on the website using carousel...")

    website_category_name = category_mapping.get(category.upper(), category.capitalize())
    category_found = False
    max_next_clicks = 3  # Increased limit for next clicks
    clicked_next_count = 0

    while not category_found and clicked_next_count <= max_next_clicks:
        try:
            # Construct XPath to find the <a> tag that contains an <h2> with the capitalized category name
            category_xpath = f"//div[@class='category-slick-content']//h2[text()='{website_category_name}']/ancestor::a"

            # Try to find the category link in the current view
            category_link = driver.find_element(By.XPATH, category_xpath)
            
            # If found, try to click it (with explicit wait for clickability)
            category_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, category_xpath))
            )

            driver.execute_script("arguments[0].scrollIntoView(true);", category_link)
            time.sleep(1) # Small pause after scrolling to ensure visibility

            category_link.click()
            print(f"Successfully selected category '{category}' on the website.")
            ui_log(f"Category '{category}' found and selected.")
            time.sleep(3) # Give some time for the new category page to load
            category_found = True
            return True

        except (NoSuchElementException, TimeoutException):
            # Category not found in the current view. Try clicking 'Next'.
            print(f"Category '{category}' not found in current view. Looking for 'Next' button...")
            ui_log(f"Category '{category}' not found, trying to click 'Next' to reveal more categories.")
            
            next_button_click_attempts = 3 # Max attempts to click the 'Next' button
            for attempt in range(next_button_click_attempts):
                try:
                    # Locate the "Next" button
                    next_button = wait.until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.slick-next.slick-arrow[aria-label='Next'][type='button']"))
                    )
                    
                    # Check if the button is disabled
                    if next_button.get_attribute("aria-disabled") == "true":
                        print("Reached the end of categories, 'Next' button is disabled.")
                        ui_log("Reached the end of categories, desired category not found.")
                        category_found = False # Ensure loop terminates
                        return False # Exit early if no more categories
                    
                    # Scroll the button into view to ensure it's actionable
                    driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                    time.sleep(1) # Small pause after scrolling
                    
                    # Get the initial transform style of the slick-track before clicking
                    slick_track = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".slick-track")))
                    initial_transform = slick_track.get_attribute("style")
                    
                    # Attempt JavaScript click
                    driver.execute_script("arguments[0].click();", next_button)
                    
                    clicked_next_count += 1
                    print(f"Clicked 'Next' button. Attempt {clicked_next_count}/{max_next_clicks}. Click retry: {attempt + 1}/{next_button_click_attempts}")
                    
                    # IMPORTANT: Wait for the 'transform' style of the slick-track to change
                    # This indicates the carousel has actually moved.
                    wait.until(lambda d: d.find_element(By.CSS_SELECTOR, ".slick-track").get_attribute("style") != initial_transform)
                    
                    time.sleep(2) # Additional brief pause after content update and carousel movement
                    break # Break from the retry loop if click was successful and content updated
                
                except (ElementClickInterceptedException, StaleElementReferenceException) as e:
                    print(f"Click interception/stale element on 'Next' button. Retrying (Attempt {attempt + 1}/{next_button_click_attempts}). Error: {e}")
                    ui_log(f"Next button click intercepted or stale. Retrying ({attempt + 1}/{next_button_click_attempts})...")
                    time.sleep(2) # Wait before retrying the click
                    if attempt == next_button_click_attempts - 1:
                        print("Max retry attempts for 'Next' button reached. Carousel did not advance.")
                        st.error("Could not click 'Next' button or carousel did not advance after multiple attempts.")
                        return False # Give up if retries fail
                except (NoSuchElementException, TimeoutException) as next_button_e:
                    print(f"No 'Next' button found or it's not clickable after previous attempts. Error: {next_button_e}")
                    st.error("No more 'Next' categories or button not found. Category not available.")
                    return False # Exit if Next button is truly not there

            else: # This else block executes if the inner for loop completes without a 'break'
                # This means the 'Next' button click attempts failed without an explicit exception
                # indicating the end of the line (e.g., if it never became clickable)
                print("Failed to click 'Next' button after all retries or carousel did not advance.")
                st.error("Failed to click 'Next' button after all retries or carousel did not advance.")
                return False # Exit function if click loop fails

        except Exception as category_e:
            print(f"An unexpected error occurred while trying to select category '{category}'. Error: {category_e}")
            st.error(f"An unexpected error occurred: {category_e}")
            return False

    if not category_found:
        print(f"Failed to select category '{category}' after trying to click 'Next' {clicked_next_count} times.")
        st.error(f"Category '{category}' not available on website after checking all visible categories.")
        return False

    return True

# Brand mapping between Excel and Website names
brand_mapping = {       
    'FLORACAL': 'FloraCal Farms',
    'NATURE\'S GRACE & WELLNESS': 'Nature\'s Grace and Wellness',
    'WANA GUMMIES': 'Wana',
    'WONDER WELLNESS': 'Wonder',
    'UPNORTH HUMBOLT': 'UpNorth',
    'LULA': 'Lula\'s',
    'JOOS': 'Joos Vapes',
    'MIDWEEK FRIDAY': 'Mid Week Friday'    
    # Continue adding brand mappings as needed.
}


# --- URL FILTER HELPERS (ADD THIS BLOCK) ---
# from urllib.parse import quote_plus

def slugify_category_for_url(excel_category: str) -> str:
    """
    Map Excel category to site category (category_mapping), then slugify for URL.
    Examples: "Pre-Rolls" -> "pre-rolls", "Flower" -> "flower"
    """
    site_cat = category_mapping.get(excel_category.upper(), excel_category)
    slug = site_cat.lower().replace("&", "and")
    slug = slug.replace("’", "'").replace("'", "")  # remove apostrophes
    slug = "-".join(slug.split())  # spaces -> hyphens
    return slug

def normalize_weight_for_url(weight: str) -> str:
    """
    Takes Excel weight and returns the URL token.
    Rules:
      - normalize like your app (g/mg, trim spaces)
      - replace '.' with '_' (e.g., 1.5g -> 1_5g)
    """
    if not weight:
        return ""
    w = normalize_weight(str(weight))  # reuses your existing normalize_weight()
    # ensure units exist (if plain number, treat as grams)
    if re.fullmatch(r"\d+(?:\.\d+)?", w):
        w = w + "g"
    return w.replace(".", "_")

def build_filtered_url(excel_category: str, brand: str, weight: str) -> str:
    """
    Build the final Terrabis URL with category (always), and optionally brand/weight.
    - Skip weight in URL for categories listed in no_weight_categories
    - Skip brand for categories in no_brand_categories
    """
    base = "https://terrabis.co/order-online/grayville/"
    params = []
    # always add category
    cat_slug = slugify_category_for_url(excel_category)
    params.append(f"dtche%5Bcategory%5D={quote_plus(cat_slug)}")
    # sort by relevance like your current flow
    params.append("dtche%5Bsortby%5D=relevance")

    site_cat = category_mapping.get(excel_category.upper(), excel_category)

    # brand (omit if category has no brand filter)
    if site_cat not in no_brand_categories and brand:
        # map brand → website display name, then lower+encode
        brand_on_site = brand_mapping.get(brand, brand).strip()
        if brand_on_site:
            params.append(f"dtche%5Bbrands%5D={quote_plus(brand_on_site.lower())}")

    # weight (omit if category has no weight filter)
    if site_cat not in no_weight_categories and weight:
        wt_token = normalize_weight_for_url(weight)
        if wt_token:
            params.append(f"dtche%5Bweight%5D={quote_plus(wt_token)}")

    return base + "?" + "&".join(params)
# --- END URL FILTER HELPERS ---


# Categories on the site that have no brand filter
no_brand_categories = ['Apparel']

def scrape_brand(brand, driver):
    """
    Function to scrape and select the brand from the search results.
    If the brand name in Excel differs from the website, it uses the mapped name.
    """
    
    # 1) Check if the brand exists in the mapping, if not, use the original name
    brand_name_on_website = brand_mapping.get(brand, brand)  # Use mapped name if exists
    
    # 2) If we're in Topicals or Accessories, skip the search box and do direct checkbox-only
    if getattr(driver, "current_category", None) in ["TOPICAL", "ACCESSORIES"]:
        labels = driver.find_elements(By.XPATH, "//label")
        for lbl in labels:
            # Normalize whitespace and match the brand name
            label_text = " ".join(lbl.text.split()).lower()
            target_text = " ".join(brand_name_on_website.lower().split())
            if target_text in label_text:
                cb = driver.find_element(By.CSS_SELECTOR, f"input[id='{lbl.get_attribute('for')}']")
                if not cb.is_selected():
                    # 🔑 scroll before clicking
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", lbl)
                    time.sleep(1)
                    cb.click()
                    print(f"✔ Selected brand (direct): {lbl.text}")
                else:
                    print(f"ℹ️ Already selected (direct): {lbl.text}")
                return True
        st.error(f"⚠️ Brand not found (direct): {brand_name_on_website}")
        return False
        
    # 3) Expand the "Brands" section if it's collapsed  
    try:
        brand_section_button = driver.find_element(By.XPATH, "//button[contains(., 'Brands')]")
        # 🔑 Always scroll to the Brands section before anything else
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", brand_section_button)
        time.sleep(1)
        if brand_section_button.get_attribute("aria-expanded") == "false":
            driver.execute_script("arguments[0].click();", brand_section_button)
            time.sleep(2)
            print("✔ Expanded the 'Brands' filter section.")

        # 🔑 Scroll again to ensure the search box area is in view
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", brand_section_button)
        time.sleep(1)

        # 3.1) Wait for the brand search box to be ready
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input.search-bar-filter__StyledInput-sc-1qap7by-1"))
        )
        search_input.clear()
        search_input.send_keys(brand_name_on_website)  # Use the mapped brand name here
        time.sleep(2)  # Allow time for the search results to appear

        # Find all available brands after searching
        brand_labels = driver.find_elements(By.XPATH, "//label")
        available_brands = [label.text.strip() for label in brand_labels]

        # 3.2) Try to find a matching brand label and select its checkbox
        for label in brand_labels:
            if brand_name_on_website.lower() in label.text.lower():  # Case-insensitive match
                brand_for_attr = label.get_attribute("for")
                brand_checkbox = driver.find_element(By.CSS_SELECTOR, f"input[id='{brand_for_attr}']")
                
                # Check if the brand checkbox is not selected, and if not, click it
                if not brand_checkbox.is_selected():
                     # 🔑 scroll before clicking
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", label)
                    time.sleep(1)
                    brand_checkbox.click()  # Click the checkbox for the brand
                    print(f"✔ Selected brand: {label.text}")
                else:
                    print(f"✔ Brand '{label.text}' is already selected.")
                
                time.sleep(2)
                return True
        
        # 3.3) If brand is not found in the available list
        print(f"⚠️ Brand '{brand_name_on_website}' is not available on this website.")
        return False

    except Exception as e:
        print(f"⚠️ Could not find or select the brand '{brand_name_on_website}'. {e}")
        return False

def normalize_weight(weight, has_unit=False):
    """
    Normalize the weight format to match the website format.
    Converts weight from GRAMS to g, and MILLIGRAMS to mg.
    """
    weight = str(weight).lower().strip()
    
    # Replace 'grams' with 'g' and 'milligrams' with 'mg'
    weight = weight.replace("grams", "g").replace("milligrams", "mg").replace("millig", "mg").replace(" ", "")
    
    # Handle cases like ".7g" instead of "0.7g"
    if weight.startswith("."):
        weight = "0" + weight  # Converts ".7g" to "0.7g"
    
    return weight

def grams_to_ounces(grams):
    """
    Converts grams to ounces and returns the fractional representation (e.g., 3.5g -> 1/8oz).
    """
    ounces = grams * 0.03527396 # Convert grams to ounces
    
    # Convert ounces to a fraction (e.g., 0.125oz -> 1/8oz)
    if ounces == 0.125:
        return '1/8oz'
    elif ounces == 0.25:
        return '1/4oz'
    elif ounces == 0.5:
        return '1/2oz'
    elif ounces == 1:
        return '1oz'
    elif ounces == 2:
        return '2oz'
    else:
        # For complex cases, round to a practical number and convert to the closest fraction
        rounded_ounces = round(ounces, 2)
        # Now we can handle small fractions that might not directly fit the exact numbers above
        if abs(rounded_ounces - 0.125) < 0.02:
            return '1/8oz'
        elif abs(rounded_ounces - 0.25) < 0.02:
            return '1/4oz'
        elif abs(rounded_ounces - 0.5) < 0.02:
            return '1/2oz'
        else:
            # For any other case (like 0.7 or 0.12oz), return the decimal representation
            return f"{rounded_ounces}oz"

def scrape_weight(weight, driver):
    """
    Selects the weight filter from the weight options.
    This function clicks the weight option based on the provided weight value.
    """
    # Normalize weight (e.g., from "0.75 GRAMS" or "1 GRAMS")
    weight_norm = normalize_weight(weight)
    # also try dropping a leading zero so "0.75g" → ".75g"
    variants = [weight_norm]
    if weight_norm.startswith("0."):
        variants.append(weight_norm[1:])

    try:
        # Find all weight filter links
        weight_links = driver.find_elements(By.CSS_SELECTOR, "a.weight__Anchor-sc-10b36p8-0.geHygR")

        # Loop through each weight option
        for link in weight_links:
            link_text = link.text.strip().lower()

            # assume pure numbers are grams (e.g. "28" -> "28g")
            if link_text.replace('.', '', 1).isdigit():
                link_text = link_text + 'g'

            # try each of our variants (with and without leading zero)
            for v in variants:
                if v in link_text:
                    link.click()
                    print(f"✔ Selected weight: {link_text}")
                    return True

        # If no matching weight found in grams, convert to ounces
        print(f"⚠️ Weight '{weight}' not found in grams. Trying to convert to ounces...")

        # Convert weight to ounces
        weight_in_ounces = grams_to_ounces(float(weight.replace('g', '').strip()))
        print(f"Converted weight: {weight_in_ounces}")

        # Search for the ounce weight
        for link in weight_links:
            link_text = link.text.strip().lower()
            if weight_in_ounces in link_text:
                link.click()  # Click the weight link
                print(f"✔ Selected weight in ounces: {link_text}")
                return True

        # If still not found
        st.error(f"⚠️ Weight '{weight_in_ounces}' not found.")
        print(f"⚠️ Weight '{weight_in_ounces}' not found.")
        return False

    except Exception as e:
        print(f"⚠️ Could not select weight '{weight}'. Error: {e}")
        return False

QUANTITY_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(pack|pk|ct|capsules|capsule|count|ea|unit|qty)", re.IGNORECASE)

from selenium.common.exceptions import WebDriverException

from selenium.common.exceptions import WebDriverException

def revive_and_reload(driver, url):
    """Tear down the dead session and reopen the URL with a fresh driver."""
    try:
        driver.quit()
    except Exception:
        pass
    new_driver, _ = get_driver()
    new_driver.get(url)
    return new_driver

def ensure_iframe_ready(driver, filtered_url, row_index):
    """
    Make the age-gate + iframe switch resilient. Retry once by recreating the driver
    if the Chrome window/session dies in-between steps.
    Returns (driver, ok_bool)
    """
    for attempt in (1, 2):
        try:
            wait = WebDriverWait(driver, 20)
            handle_age_verification_popup(driver, wait)

            time.sleep(2)  # brief settle
            iframe = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, "dutchie--embed__iframe"))
            )
            driver.switch_to.frame(iframe)

            WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-testid='product-list-item']"))
            )
            time.sleep(2)
            return driver, True

        except (TimeoutException, WebDriverException) as e:
            if attempt == 2:
                # after one retry, give up on this row
                st.error(f"Iframe / age-gate failed (row {row_index}). Error: {e}")
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                return driver, False

            # revive & retry once
            ui_log(f"Recovering Chrome session for row {row_index} and retrying once...")
            driver = revive_and_reload(driver, filtered_url)

    return driver, False


def safe_get(driver, url, retry=False):
    """
    Navigate to URL. If the window/session is gone, optionally recreate the driver once.
    Returns (driver, ok_bool). If driver was recreated, the new driver is returned.
    """
    try:
        driver.get(url)
        return driver, True
    except WebDriverException as e:
        msg = str(e).lower()
        if (("no such window" in msg or "web view not found" in msg or "disconnected" in msg) and not retry):
            # Recreate driver once
            print("Driver/session lost; recreating driver and retrying url once...")
            try:
                driver.quit()
            except Exception:
                pass
            new_driver, _ = get_driver()
            new_driver.get(url)
            return new_driver, True
        return driver, False


def extract_and_normalize_quantity(text):
    """
    Extracts numerical quantity and normalizes the unit.
    Returns (quantity_number, normalized_unit_type) or (None, None).
    Normalized unit types: 'mg', 'ml', 'count'
    """
    match = QUANTITY_RE.search(text)
    if match:
        number = float(match.group(1))
        unit = match.group(2).lower()
        if unit in ['pack', 'pk', 'ct', 'capsules', 'capsule', 'count', 'ea', 'unit', 'qty']:
            return number, 'count'
    return None, None

RATIO_RE = re.compile(
    r"""
    # Optional opening parenthesis or bracket, followed by optional whitespace
    (?:[\(\[]?\s*)?
    (?:
        # Pattern A: Cannabinoid(s) first, then numerical ratio (e.g., THC:CBD 1:2, THC/CBD/CBG 1:1:1)
        (?:THC|CBD|CBG|CBN|CBC)(?:[\s\/:]?(?:THC|CBD|CBG|CBN|CBC))* \s* \d+[:\/]\d+(?:[:\/]\d+)*
        |
        # Pattern B: Numerical ratio first, then Cannabinoid(s) (e.g., 1:1 THC:CBG, 8:1 CBD/THC)
        \d+[:\/]\d+(?:[:\/]\d+)* \s* (?:THC|CBD|CBG|CBN|CBC)(?:[\s\/:]?(?:THC|CBD|CBG|CBN|CBC))*
    )
    # Optional whitespace followed by optional closing parenthesis or bracket
    (?:\s*[\)\]]?)?
    """,
    re.VERBOSE | re.IGNORECASE
)

def extract_flavors(text, flavor_list):
    """
    Extracts flavors from a given text based on a predefined list of flavors.
    Returns a list of unique flavors found.
    """
    found_flavors = set()
    text_lower = text.lower()
    
    # Sort flavors by length descending to match longer phrases first (e.g., "blue raspberry" before "blue")
    sorted_flavors = sorted(flavor_list, key=len, reverse=True)

    for flavor in sorted_flavors:
        # Use word boundaries to avoid partial matches (e.g., 'grape' in 'grapefruit')
        if re.search(r'\b' + re.escape(flavor) + r'\b', text_lower):
            found_flavors.add(flavor)
    return list(found_flavors)

def word_match_score(a, b):
    """
    Returns fraction of words in a that also appear in b.
    """
    wa = re.findall(r"\w+", a.lower())
    wb = set(re.findall(r"\w+", b.lower()))
    if not wa:
        return 0
    matches = sum(1 for w in wa if w in wb)
    return matches / len(wa)

# ---- Page config + polished theme ----
st.set_page_config(page_title="Terrabis Scraper", page_icon="🌿", layout="wide")

st.markdown("""
<style>
header[data-testid="stHeader"] { background: transparent; }
#MainMenu, footer {visibility: hidden;}
:root { --brand:#6a2af1; --soft:#F5F6FA; --ok:#12B886; --muted:#6B7280; }
.block-container { padding-top: 1.2rem; }
.topbar { background:white; border:1px solid #EEF0F4; border-radius:12px; padding:14px 18px; margin-bottom:12px; }
.brand { font-weight:700; color:var(--brand); font-size:1.15rem; letter-spacing:.2px;}
.subtle { color: var(--muted); }
.status { display:inline-block; padding:4px 10px; border-radius:999px; font-size:.82rem; border:1px solid #E5E7EB; }
.status.running { background:#ECFDF5; color:var(--ok); border-color:#C7F9E9; }
.status.idle { background:#F3F4F6; color:#6B7280; }
.kpi { background:white; border:1px solid #EEF0F4; border-radius:12px; padding:12px 14px; height:92px; }
.kpi .label { color:var(--muted); font-size:.82rem; }
.kpi .value { font-weight:700; font-size:1.4rem; margin-top:6px; }
.footer { margin-top:10px; padding:8px 12px; border-radius:10px; background:var(--soft); color:#4B5563; font-size:.9rem; border:1px solid #EEF0F4; }
</style>
""", unsafe_allow_html=True)

# ---- UI session state (UI-only; scraping logic unchanged) ----
if "running" not in st.session_state: st.session_state.running = False
if "rows_done" not in st.session_state: st.session_state.rows_done = 0
if "rows_matched" not in st.session_state: st.session_state.rows_matched = 0
if "log_lines" not in st.session_state: st.session_state.log_lines = []
if "stop_requested" not in st.session_state: st.session_state.stop_requested = False

def render_status(box):
    cls = "running" if st.session_state.running else "idle"
    txt = "Running" if st.session_state.running else "Idle"
    box.markdown(f"<span class='status {cls}'>{txt}</span>", unsafe_allow_html=True)

def render_kpis(kpi_rows, kpi_match, kpi_state):
    kpi_rows.markdown(f"<div class='kpi'><div class='label'>Rows processed</div>"
                      f"<div class='value'>{st.session_state.rows_done}</div></div>", unsafe_allow_html=True)
    kpi_match.markdown(f"<div class='kpi'><div class='label'>Matches saved</div>"
                       f"<div class='value'>{st.session_state.rows_matched}</div></div>", unsafe_allow_html=True)
    kpi_state.markdown(f"<div class='kpi'><div class='label'>Status</div>"
                       f"<div class='value'>{'OK' if st.session_state.rows_done else '-'}</div></div>", unsafe_allow_html=True)

# ---- Top bar (brand + status + KPIs) ----
# ---- Top bar (brand + status only; KPIs will live in Overview tab) ----
with st.container():
    st.markdown('<div class="topbar">', unsafe_allow_html=True)
    _cols = st.columns([3,2])
    with _cols[0]:
        st.markdown("**<span class='brand'>Terrabis Scraper</span>**<br>"
                    "<span class='subtle'>URL-filtered, Excel-updated</span>", unsafe_allow_html=True)
    status_box = _cols[1].empty()
    st.markdown("</div>", unsafe_allow_html=True)

render_status(status_box)
# (no render_kpis() here anymore)


# ---- Tabs & placeholders (Overview / Logs / Settings) ----
tabs = st.tabs(["Overview", "Logs", "Settings"])
# ---- Overview tab (KPIs + progress + preview) ----
with tabs[0]:
    # KPI tiles (Overview-owned)
    ov1, ov2, ov3 = st.columns(3)
    ov_rows  = ov1.empty()
    ov_match = ov2.empty()
    ov_state = ov3.empty()

    def render_overview_kpis():
        ov_rows.markdown(
            f"<div class='kpi'><div class='label'>Rows processed</div>"
            f"<div class='value'>{st.session_state.rows_done}</div></div>", unsafe_allow_html=True)
        ov_match.markdown(
            f"<div class='kpi'><div class='label'>Matches saved</div>"
            f"<div class='value'>{st.session_state.rows_matched}</div></div>", unsafe_allow_html=True)
        ov_state.markdown(
            f"<div class='kpi'><div class='label'>Status</div>"
            f"<div class='value'>{'OK' if st.session_state.rows_done else '-'}</div></div>", unsafe_allow_html=True)

    render_overview_kpis()

    # Download button placeholder (we’ll fill when done)
    download_placeholder = st.empty()
    st.markdown("")

    # Progress bar
    progress = st.progress(0, text="Idle")

    # Preview placeholder (updated as we go / at end)
    st.caption("Preview (first 50 rows)")
    preview_placeholder = st.empty()

with tabs[1]:
    logs_placeholder = st.empty()
    logs_placeholder.code("—")
    # Central logger that writes ONLY to the Logs tab
    # --- Log helper (REPLACE your existing ui_log with this) ---
    def ui_log(*parts):
        msg = " ".join(str(p) for p in parts)
        st.session_state.log_lines.append(msg)
        # render last ~200 lines into the Logs tab box
        if 'logs_placeholder' in globals():
            logs_placeholder.code("\n".join(st.session_state.log_lines[-200:]) or "—")

with tabs[2]:
    st.caption("Tuning")
    st.write("- Sort: relevance")
    st.write("- Weight omitted in URL for: Edibles / Topicals / Accessories / Apparel")
    st.write("- Brand & category mapping per code")
    st.write("- Headless Chrome via undetected_chromedriver")

# Streamlit App Interface
# ---- Sidebar (mimic target UI) ----
with st.sidebar:
    st.markdown("### 📄 Input")
    with st.expander("Instructions"):
        st.markdown("""
        1. **Upload an Excel file** with the *Pricing Research* sheet.
        2. Choose the category and click **Start**.
        3. The scraper will update the workbook in-memory.
        4. When done, click **Download** to save the updated file.
        """)

    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx'])
    st.caption("Sheet: **Pricing Research**")

    c1, c2 = st.columns(2)
    btn_start = c1.button(
        "Start", type="primary", use_container_width=True,
        disabled=st.session_state.running or uploaded_file is None
    )
    btn_stop  = c2.button(
        "Stop", use_container_width=True,
        disabled=not st.session_state.running
    )
    if btn_stop:
        st.session_state.stop_requested = True


if uploaded_file:
    # Load the file into memory buffer *once* when uploaded
    save_updated_excel_to_memory(uploaded_file) 
    
    # Load the file from the BUFFER
    excel_buffer.seek(0)
    df = pd.read_excel(excel_buffer, sheet_name="Pricing Research")
    
    # Extract unique categories and brands from the 'Category' and 'Brand' columns
    categories = df['Category'].unique()
    brands = df['Brand'].unique()

    # Dropdown in Sidebar for Category Selection
    selected_category = st.sidebar.selectbox("Select Category to Scrap", categories)
    
    # Filter the data based on the selected category
    filtered_data = df[df['Category'] == selected_category]

    # Show the number of products in the selected category
    num_products = len(filtered_data)
    ui_log('-------------------------------------------------------------------------')
    ui_log(f"Number of products in the '{selected_category}' category: {num_products}")

    # Initial preview for the selected category
    try:
        preview_placeholder.dataframe(filtered_data.head(50), use_container_width=True, height=410)
    except Exception:
        pass

    # Prime the progress bar
    progress.progress(
        0,
        text=f"Ready — {num_products} rows in '{selected_category}'"
    )

    # Button to Start Scraping
    # Button to Start Scraping
    if 'btn_start' in globals() and btn_start:
        st.session_state.running = True
        st.session_state.stop_requested = False
        st.session_state.rows_done = 0
        st.session_state.rows_matched = 0
        render_status(status_box)
        render_overview_kpis()

        ui_log("Scraping started for category:", selected_category)
        # ui_log("Scraping started for category:", selected_category)

        # Initialize the driver
        driver, wait = get_driver()

        # --- LOOP ROWS IN SELECTED CATEGORY ---
        for row_index, row in filtered_data.iterrows():

            # Stop button support
            if st.session_state.stop_requested:
                ui_log("Stop requested — finishing after current row.")
                break

            # --- URL-BASED FILTERING (REPLACES brand/weight clicks) ---
            try:
                # 1) Row fields
                selected_category_row = row['Category']  # keep separate from sidebar var
                brand = str(row['Brand']) if pd.notna(row['Brand']) else ""
                weight = str(row['Weight']) if pd.notna(row['Weight']) else ""

                # 2) Build filtered URL (category always; brand/weight when allowed)
                filtered_url = build_filtered_url(selected_category_row, brand, weight)
                ui_log(f"[Row {row_index}] Navigating with URL filters → {filtered_url}")
                print(f"[Row {row_index}] GET {filtered_url}")
                driver, ok = safe_get(driver, filtered_url)
                if not ok:
                    st.error(f"Could not open page for row {row_index}. Skipping.")
                    save_data_to_file(row_index, " ", " ", " ", " ")
                    continue


                # 3) Age gate (if it appears)
                # 3–4) Age-gate + iframe (self-healing)
                driver, ok = ensure_iframe_ready(driver, filtered_url, row_index)
                if not ok:
                    save_data_to_file(row_index, " ", " ", " ", " ")
                    continue
                ui_log(f"Switched to Dutchie iframe for row {row_index} (URL-filtered).")


                # 5) Prepare tokens (used by your existing matching logic below)
                #    (note: keep names identical to what your matching block expects)
                selected_category = selected_category_row
                raw_brand = TOKEN_RE.findall(brand)
                brand_tokens = {t.replace(" ", "").lower() for t in raw_brand}
                website_cat      = category_mapping.get(selected_category, selected_category)
                raw_cat          = TOKEN_RE.findall(website_cat)
                collapsed_cat    = [t.replace(" ", "") for t in raw_cat]
                category_tokens  = set()
                for tok in collapsed_cat:
                    lc = tok.lower()
                    category_tokens.add(lc)
                    if lc.endswith("s"):
                        category_tokens.add(lc[:-1])
                # normalized_weight is still needed later for enforcement in no-weight categories
                normalized_weight = normalize_weight(weight)

            except Exception as nav_e:
                st.error(f"URL navigation error (row {row_index}): {nav_e}")
                save_data_to_file(row_index, " ", " ", " ", " ")
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                continue
            # --- END URL-BASED FILTERING ---

            # --- PRODUCT MATCHING START ---
            try:
                # wait for the product tiles to appear
                WebDriverWait(driver, 8).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-testid='product-list-item']"))
                )
                time.sleep(5)  # give a bit extra for everything to render

                # scrape all product names on the page
                product_tiles = driver.find_elements(By.CSS_SELECTOR, "div[data-testid='product-list-item']")
                scraped = []

                # grab the target from Excel
                target_name = row['Product Name']
                # Extract and normalize quantity from Excel product name
                excel_qty_num, excel_qty_unit = extract_and_normalize_quantity(target_name)

                # Identify all parts of quantity strings from the Excel product name
                excel_quantity_parts_to_exclude = set()
                for m in QUANTITY_RE.finditer(target_name):
                    full_match_str = m.group(0).lower() # e.g., "10 pk", "100mg"
                    # Tokenize the full match string using TOKEN_RE to get its constituent tokens
                    # and add them to the set of parts to exclude
                    for part_token in TOKEN_RE.findall(full_match_str):
                        excel_quantity_parts_to_exclude.add(part_token.replace(" ", "").lower())

                # Identify all parts of ratio strings from the Excel product name
                excel_ratio_parts_to_exclude = set()
                for m in RATIO_RE.finditer(target_name):
                    full_match_str = m.group(0).lower()
                    for part_token in TOKEN_RE.findall(full_match_str):
                        excel_ratio_parts_to_exclude.add(part_token.replace(" ", "").lower())
                
                # Extract flavors from Excel product name (still needed for strict matching)
                excel_flavors = extract_flavors(target_name, FLAVOR_LIST)
                excel_flavor_tokens = set(excel_flavors) # Store as set for comparison

                # tokenize and collapse spaces in units (so "3.5 g" → "3.5g")
                raw_wa       = TOKEN_RE.findall(target_name)
                collapsed_wa = [t.replace(" ", "") for t in raw_wa]

                # extract weight tokens (anything starting with a digit)
                excel_weight_tokens = [
                    t for t in collapsed_wa
                    if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', t.lower())
                ]

                # Filter Excel tokens for keywords. Flavors are *not* excluded here.
                excel_keyword_tokens_list = []
                for t in collapsed_wa:
                    normalized_t_lower = t.lower()
                    # Check if it's a weight token
                    if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', normalized_t_lower):
                        continue
                    # Check if the token is part of an identified quantity
                    if normalized_t_lower in excel_quantity_parts_to_exclude:
                        continue
                    # Check if the token is part of an identified ratio
                    if normalized_t_lower in excel_ratio_parts_to_exclude:
                        continue
                    # Check if it's a brand/category/stopword token
                    if normalized_t_lower in brand_tokens or \
                       normalized_t_lower in STOPWORDS or \
                       normalized_t_lower in category_tokens or \
                       (normalized_t_lower.endswith("s") and normalized_t_lower[:-1] in category_tokens):
                        continue
                    # Flavors are NO LONGER EXCLUDED HERE; they will contribute to the general score.
                    excel_keyword_tokens_list.append(normalized_t_lower)

                # Convert to set for efficient lookup during comparison
                excel_keyword_tokens_set = set(excel_keyword_tokens_list)
                excel_tokens_display = [t.title() for t in excel_keyword_tokens_list]

                ui_log(f"🔎 **Product name:** {target_name}")
                print(f"⚖️ **Excel weight tokens:** {', '.join(excel_weight_tokens)}")
                print(f"📦 **Excel quantity:** {excel_qty_num} {excel_qty_unit if excel_qty_unit else 'N/A'}")
                print(f"🎨 **Excel flavors:** {', '.join(excel_flavors) if excel_flavors else 'N/A'}") # Display extracted flavors
                print(f"🔍 **Excel tokens (cleaned):** {', '.join(excel_tokens_display)}") # Now truly cleaned

                for tile in product_tiles:
                    name = tile.find_element(By.CSS_SELECTOR, "div.full-card__Name-sc-11z5u35-4").text
                    
                    # Extract URL
                    product_url = " "
                    try:
                        # Attempt to find the anchor tag for the product URL
                        url_element = tile.find_element(By.TAG_NAME, "a")
                        product_url = url_element.get_attribute("href")
                    except NoSuchElementException:
                        print(f"⚠️ URL not found for product '{name}'")

                    # Extract THC
                    thc_content = " "
                    try:
                        thc_element = tile.find_element(By.CSS_SELECTOR, "div.full-card__Potency-sc-11z5u35-8 > div")
                        raw_thc_text = thc_element.text
                        thc_content = clean_thc_value(raw_thc_text) # Apply the cleaning function here
                    except NoSuchElementException:
                        pass # THC might not be present for all products

                    # Extract Price
                    discounted_price = " "
                    original_price = " "
                    try:
                        option_tile_button = tile.find_element(By.CSS_SELECTOR, "button[data-testid='option-tile']")
                        
                        # First, try to find the original price span, as its presence dictates the logic
                        try:
                            original_price_element_if_discount = option_tile_button.find_element(By.CSS_SELECTOR, "span.optionstyles__OriginalPrice-sc-vu6uvs-2")
                            # If this element is found, it means there's a discount
                            original_price = original_price_element_if_discount.text
                            # The 'b' tag then holds the discounted price
                            discounted_price_element = option_tile_button.find_element(By.TAG_NAME, "b")
                            discounted_price = discounted_price_element.text
                        except NoSuchElementException:
                            # If original_price_element_if_discount is NOT found, it means no discount
                            # In this case, the 'b' tag holds the original price
                            single_price_element = option_tile_button.find_element(By.TAG_NAME, "b")
                            original_price = single_price_element.text
                            discounted_price = " " # As per new requirement
                            
                    except NoSuchElementException:
                        print(f"⚠️ Price information not found for product '{name}'")

                    # Extract and normalize quantity from Site product name
                    site_qty_num, site_qty_unit = extract_and_normalize_quantity(name)

                    # Identify all parts of quantity strings from the Site product name
                    site_quantity_parts_to_exclude = set()
                    for m in QUANTITY_RE.finditer(name):
                        full_match_str = m.group(0).lower()
                        for part_token in TOKEN_RE.findall(full_match_str):
                            site_quantity_parts_to_exclude.add(part_token.replace(" ", "").lower())

                    # Identify all parts of ratio strings from the Site product name
                    site_ratio_parts_to_exclude = set()
                    for m in RATIO_RE.finditer(name):
                        full_match_str = m.group(0).lower()
                        for part_token in TOKEN_RE.findall(full_match_str):
                            site_ratio_parts_to_exclude.add(part_token.replace(" ", "").lower())
                    
                    # Extract flavors from Site product name
                    site_flavors = extract_flavors(name, FLAVOR_LIST)
                    site_flavor_tokens = set(site_flavors) # Store as set for comparison

                    raw_wb = TOKEN_RE.findall(name)
                    collapsed_wb = [t.replace(" ", "") for t in raw_wb]

                    site_weight_tokens = [
                        t for t in collapsed_wb
                        if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', t.lower())
                    ]

                    # Filter Site tokens for keywords. Flavors are *not* excluded here.
                    site_keyword_tokens_list = []
                    for t in collapsed_wb:
                        normalized_t_lower = t.lower()
                        # Check if it's a weight token
                        if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', normalized_t_lower):
                            continue
                        # Check if it's part of an identified quantity
                        if normalized_t_lower in site_quantity_parts_to_exclude:
                            continue
                        # Check if the token is part of an identified ratio
                        if normalized_t_lower in site_ratio_parts_to_exclude:
                            continue
                        # Check if it's a brand/category/stopword token
                        if normalized_t_lower in brand_tokens or \
                           normalized_t_lower in STOPWORDS or \
                           normalized_t_lower in category_tokens or \
                           (normalized_t_lower.endswith("s") and normalized_t_lower[:-1] in category_tokens):
                            continue
                        # Flavors are NO LONGER EXCLUDED HERE; they will contribute to the general score.
                        site_keyword_tokens_list.append(normalized_t_lower)

                    # Add site_qty_num, site_qty_unit, and site_keyword_tokens_list to scraped tuple
                    scraped.append((name, product_url, discounted_price, original_price, thc_content, site_weight_tokens, site_qty_num, site_qty_unit, site_flavors, site_keyword_tokens_list))

            except TimeoutException:
                ui_log(f"⚠️ No products found for brand **{brand}** in category **{selected_category}**")
                # When no products are found, ensure N/A is written to Excel
                save_data_to_file(row_index, " ", " ", " ", " ") # Save "N/A" for this row
                continue
            except Exception as e:
                st.error(f"An error occurred while scraping product tiles: {e}")
                save_data_to_file(row_index, " ", " ", " ", " ")
                continue

            # Initialize lists to store multiple matches if fuzzy matching
            matched_urls = []
            matched_discounted_prices = []
            matched_original_prices = []
            matched_thc_contents = []

            best_match, best_score = None, 0.0

            if len(excel_keyword_tokens_set) <= 3:
                match_threshold = 0.6  # 60%
                print("Threshold set to 60% due to <= 3 Excel tokens.")
            else:
                match_threshold = 0.75 # 75%
                print("Threshold set to 75% due to > 3 Excel tokens.")

            # Update the loop to unpack new scraped fields
            for name, url, discounted_price, original_price, thc_content, site_weight_tokens, site_qty_num, site_qty_unit, site_flavors, site_keyword_tokens_list in scraped:
                # --- QUANTITY COMPARISON LOGIC ---
                quantity_match = True
                if excel_qty_num is not None and site_qty_num is not None:
                    if not (excel_qty_num == site_qty_num and excel_qty_unit == site_qty_unit):
                        quantity_match = False
                        print(f"  Quantity mismatch: Excel '{excel_qty_num} {excel_qty_unit}' vs Site '{site_qty_num} {site_qty_unit}' for '{name}'")

                # if no‐weight category, enforce exact weight match before comparing
                weight_enforced_match = True
                if website_cat in no_weight_categories:
                    # Normalize site_weight_tokens for consistent comparison (e.g., '1g', '500mg')
                    normalized_site_weight_tokens = [normalize_weight(swt) for swt in site_weight_tokens]

                    # Now, compare the normalized weight from the Excel 'Weight' column
                    # (which is 'normalized_weight') against the site's normalized weight tokens.
                    if normalized_weight not in normalized_site_weight_tokens:
                        weight_enforced_match = False
                        # Update print statement to show the actual Excel column weight being used
                        print(f"  Weight mismatch for no-weight category: Excel '{normalized_weight}' vs Site '{', '.join(site_weight_tokens)}' for '{name}'")

                # Flavor Matching Logic
                flavor_match = True
                if excel_flavors: # If there are flavors in the Excel product name
                    # Check if ALL Excel flavors are present in the site product's flavors
                    if not all(f in site_flavor_tokens for f in excel_flavor_tokens):
                        flavor_match = False
                        print(f"  Flavor mismatch: Excel '{', '.join(excel_flavors)}' vs Site '{', '.join(site_flavors) if site_flavors else 'N/A'}' for '{name}'")
                # If Excel product has no specific flavors, then any flavor on site is acceptable.
                # The 'flavor_match' remains True in this case, meaning no flavor mismatch prevents a match.

                # Convert to set for efficient comparison
                lc_site_keyword_tokens_set = set(site_keyword_tokens_list)
                site_tokens_display = [t.title() for t in site_keyword_tokens_list] # For display

                print(f"⚖️ **Site weight tokens for “{name}”:** {', '.join(site_weight_tokens)}")
                print(f"📦 **Site quantity for “{name}”:** {site_qty_num} {site_qty_unit if site_qty_unit else 'N/A'}")
                print(f"🎨 **Site flavors for “{name}”:** {', '.join(site_flavors) if site_flavors else 'N/A'}") # Display extracted site flavors
                print(f"👁️ **Site tokens for “{name}” (cleaned):** {', '.join(site_tokens_display)}")
                print(f"💰 **Site Price for “{name}”:** Discounted: {discounted_price}, Original: {original_price}")
                print(f"🌿 **Site THC for “{name}”:** {thc_content}")
                print(f"🌐 **Site URL for “{name}”:** {url}")

                # compare on lowercase using the cleaned keyword token sets
                common = [w for w in excel_keyword_tokens_set if w in lc_site_keyword_tokens_set]
                common_tokens_display = [t.title() for t in common]
                print(f"🔗 **Common tokens:** {', '.join(common_tokens_display)}")

                # compute score based only on keyword tokens
                if not excel_keyword_tokens_set:
                    score = 0.0
                else:
                    score = len(common) / len(excel_keyword_tokens_set)
                print(f"      Score for “{name}”: {score:.0%}")

                # If quantity matched (or wasn't applicable for strict match) and keyword score is good
                # And now, ensure weight also matched if it's a no_weight_category
                # Ensure flavors also matched if present in Excel
                if quantity_match and weight_enforced_match and flavor_match and score >= match_threshold:
                    # Instead of just taking the best, collect all valid matches
                    # This is for the scenario where multiple products could fuzzy match
                    matched_urls.append(url)
                    matched_discounted_prices.append(discounted_price)
                    matched_original_prices.append(original_price)
                    matched_thc_contents.append(thc_content)

                    # Update best_match for display purposes if a higher score is found
                    if score > best_score:
                        best_match_name = name
                        best_score = score

            # if matched_urls: # If any matches were found
            if matched_urls:  # If any matches were found
                ui_log(f"✅ Matched “{target_name}” → “{best_match_name}” ({best_score:.0%})")
                ui_log(f"   **URL(s):** {', '.join(matched_urls)}")
                ui_log(f"   **Price(s):** Discounted: {', '.join(map(str, matched_discounted_prices))} (Original: {', '.join(map(str, matched_original_prices))})")
                ui_log(f"   **THC(s):** {', '.join(matched_thc_contents)}")
                save_data_to_file(row_index, matched_discounted_prices, matched_original_prices, matched_thc_contents, matched_urls)
            else:
                ui_log(f"⚠️ No ≥{int(match_threshold * 100)}% match for “{target_name}”.")
                save_data_to_file(row_index, " ", " ", " ", " ")

            # ⬇️ PASTE THE KPI/PROGRESS/LOGS SNIPPET HERE ⬇️
            # --- KPI + progress + logs
            st.session_state.rows_done += 1
            if 'matched_urls' in locals() and matched_urls:
                st.session_state.rows_matched += 1

            pct = int((st.session_state.rows_done / max(1, num_products)) * 100)
            progress.progress(pct, text=f"Processed {st.session_state.rows_done}/{num_products}")

            # refresh top KPIs
            render_overview_kpis()

            # lightweight log line
            msg = f"Row {row_index}: {'matched' if ('matched_urls' in locals() and matched_urls) else 'no match'}"
            st.session_state.log_lines.append(msg)
            logs_placeholder.code("\n".join(st.session_state.log_lines[-200:]) or "—")

            # --- PRODUCT MATCHING END ---
            time.sleep(3)

        # ui_log("Scraping completed for category:", selected_category)
        # After all rows
        ui_log("Scraping completed for category:", selected_category)
        st.session_state.running = False
        render_status(status_box)
        render_overview_kpis()

        done_text = "Finished" if not st.session_state.stop_requested else "Stopped early"
        progress.progress(100, text=f"{done_text} — rows: {st.session_state.rows_done}")
                
        driver.quit()
        if excel_buffer is not None:
            st.download_button(
                label="Download Updated Excel File",
                data=excel_buffer.getvalue(),
                file_name=f"updated_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# # --- distutils shim for Python 3.12+ (must be BEFORE importing undetected_chromedriver) ---
# try:
#     from distutils.version import LooseVersion  # noqa: F401
# except Exception:
#     import types, sys, re
#     from packaging.version import Version as _PV

#     def _parts_list(v: str):
#         return re.split(r"[^\w]+", v.strip())  # split on dots/dashes, keep digits/letters

#     class LooseVersion(str):
#         # Minimal drop-in: has `.version` (list of ints/strs) and `.vstring` (str), supports comparisons
#         def __new__(cls, v):
#             obj = str.__new__(cls, v)
#             parts = [p for p in _parts_list(str(v)) if p != ""]
#             def _cast(x):
#                 try:
#                     return int(x)
#                 except Exception:
#                     return x
#             obj.version = [_cast(p) for p in parts]
#             return obj

#         @property
#         def vstring(self):
#             return str(self)

#         def _v(self):
#             return _PV(str(self))
#         def __lt__(self, o): return self._v() <  _PV(str(o))
#         def __le__(self, o): return self._v() <= _PV(str(o))
#         def __gt__(self, o): return self._v() >  _PV(str(o))
#         def __ge__(self, o): return self._v() >= _PV(str(o))
#         def __eq__(self, o): return self._v() == _PV(str(o))

#     dv = types.ModuleType("distutils")
#     dv_version = types.ModuleType("distutils.version")
#     dv_version.LooseVersion = LooseVersion
#     dv.version = dv_version
#     sys.modules["distutils"] = dv
#     sys.modules["distutils.version"] = dv_version
# # ----------------------------------------------------------------------





# import re
# import time
# import io
# import streamlit as st
# import pandas as pd
# import undetected_chromedriver as uc
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.service import Service
# from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains
# from openpyxl import load_workbook 
# from openpyxl.utils import get_column_letter # Add this import
# from selenium.common.exceptions import WebDriverException
# from selenium.webdriver.support import expected_conditions as EC
# import os, subprocess, re
# from urllib.parse import urlencode
# # import os, subprocess, re

# def _find_chrome_binary():
#     env = os.environ.get("UC_CHROME_BINARY")
#     if env and os.path.exists(env):
#         return env
#     for p in ("/usr/bin/chromium",
#               "/usr/bin/chromium-browser",
#               "/usr/bin/google-chrome",
#               "/usr/bin/google-chrome-stable"):
#         if os.path.exists(p):
#             return p
#     return None

# def _chrome_major(chrome_bin):
#     try:
#         out = subprocess.check_output([chrome_bin, "--version"]).decode()
#         m = re.search(r"(\d+)\.", out)
#         return int(m.group(1)) if m else 120
#     except Exception:
#         return 120
# # -----------------------------------


# # ---------- Headless helpers ----------
# from selenium.common.exceptions import WebDriverException

# def stable_click(driver, elem):
#     """Scroll into view, try normal click, fallback to JS click (headless-safe)."""
#     driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", elem)
#     time.sleep(0.2)
#     try:
#         elem.click()
#         return True
#     except WebDriverException:
#         try:
#             driver.execute_script("arguments[0].click();", elem)
#             return True
#         except WebDriverException:
#             return False

# def wait_visible(driver, locator, timeout=20):
#     return WebDriverWait(driver, timeout).until(EC.visibility_of_element_located(locator))

# def wait_present(driver, locator, timeout=20):
#     return WebDriverWait(driver, timeout).until(EC.presence_of_element_located(locator))
# # --------------------------------------

# def type_react_input(driver, el, text):
#     """
#     Sets value on a React-controlled <input> and dispatches the right events so filtering happens in headless too.
#     """
#     driver.execute_script("""
#     const el = arguments[0], val = arguments[1];
#     const proto = el.__proto__ || HTMLInputElement.prototype;
#     const setter = Object.getOwnPropertyDescriptor(proto, 'value').set
#                 || Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
#     setter.call(el, val);
#     el.dispatchEvent(new Event('input',  {bubbles: true}));
#     el.dispatchEvent(new Event('change', {bubbles: true}));
#     """, el, text)

# def click_show_more_until_exhausted(driver):
#     """
#     Some filters hide brands behind a 'Show more' button.
#     Click it repeatedly if present.
#     """
#     import time
#     while True:
#         try:
#             btn = driver.find_element(By.XPATH, "//button[contains(., 'Show more') or contains(., 'More')]")
#             if btn.is_displayed() and btn.is_enabled():
#                 driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
#                 time.sleep(0.2)
#                 stable_click(driver, btn)
#                 time.sleep(0.6)
#             else:
#                 break
#         except NoSuchElementException:
#             break

# def scroll_filter_panel_to_find_label(driver, brand_text, max_scrolls=30):
#     """
#     When the brand list is virtualized, we need to scroll the panel to load more labels.
#     """
#     import time
#     # The list container is usually the element right after the 'Brands' button
#     try:
#         panel = driver.find_element(By.XPATH, "//button[contains(., 'Brands')]/following-sibling::*[1]")
#     except NoSuchElementException:
#         panel = None

#     brand_norm = " ".join(brand_text.lower().split())

#     for _ in range(max_scrolls):
#         # Try to find the label in the currently loaded chunk
#         labels = driver.find_elements(By.XPATH, "//label")
#         for lbl in labels:
#             t = " ".join(lbl.text.lower().split())
#             if brand_norm in t and lbl.is_displayed():
#                 stable_click(driver, lbl)
#                 return True
#         # If we can scroll a dedicated panel, do that; otherwise scroll window
#         if panel:
#             driver.execute_script("arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].clientHeight;", panel)
#         else:
#             driver.execute_script("window.scrollBy(0, 600);")
#         time.sleep(0.4)

#     return False

# # ---- URL helpers to pre-filter brand/category on Terrabis ----
# category_slug_map = {
#     "Edibles": "edibles",
#     "Flower": "flower",
#     "Vaporizers": "vaporizers",
#     "Concentrates": "concentrates",
#     "Topicals": "topicals",
#     "Pre-Rolls": "pre-rolls",
#     "Tinctures": "tinctures",
#     "Apparel": "apparel",
#     "Accessories": "accessories",
# }

# def slugify_brand_for_param(name: str | None) -> str | None:
#     if not name:
#         return None
#     s = name.lower()
#     s = s.replace("&", " and ")
#     s = s.replace("’", "").replace("'", "").replace("`", "")
#     s = re.sub(r"[^a-z0-9]+", "-", s)
#     s = re.sub(r"-+", "-", s).strip("-")
#     return s

# def build_terrabis_url(city_slug: str, category_site_name: str, brand_site_name: str | None) -> str:
#     cat_slug = category_slug_map.get(category_site_name, category_site_name.lower())
#     base = f"https://terrabis.co/order-online/{city_slug}/"
#     params = {
#         "dtche[category]": cat_slug,
#         "dtche[sortby]": "relevance",
#     }
#     bslug = slugify_brand_for_param(brand_site_name)
#     if bslug:  # include brand only when provided
#         params["dtche[brands]"] = bslug
#     return f"{base}?{urlencode(params)}"

# def open_terrabis_with_brand(driver, wait, city_slug: str, category_site_name: str, brand_site_name: str | None, row_index: int) -> bool:
#     """
#     Navigate to Terrabis with category (+ optional brand) applied via query params.
#     Stay on Terrabis, switch into the Dutchie iframe, and wait for product tiles.
#     """
#     url = build_terrabis_url(city_slug, category_site_name, brand_site_name)
#     driver.switch_to.default_content()
#     driver.get(url)

#     # Close age gate on host page
#     try:
#         handle_age_verification_popup(driver, wait)
#     except Exception:
#         pass

#     try:
#         iframe = WebDriverWait(driver, 25).until(
#             EC.presence_of_element_located((
#                 By.CSS_SELECTOR,
#                 "iframe#dutchie--embed__iframe, iframe[id*='dutchie'], iframe[src*='dutchie.com/embedded-menu']"
#             ))
#         )
#         WebDriverWait(driver, 25).until(EC.frame_to_be_available_and_switch_to_it(iframe))
#         st.info(f"Switched to Dutchie iframe for row {row_index}.")

#         # settle + optional cookie
#         try:
#             WebDriverWait(driver, 10).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
#         except Exception:
#             pass
#         try:
#             cookie_btn = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((
#                 By.XPATH, "//button[normalize-space()='Accept' or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept')]"
#             )))
#             stable_click(driver, cookie_btn)
#             time.sleep(0.4)
#         except Exception:
#             pass

#         # wait for tiles
#         WebDriverWait(driver, 35).until(EC.any_of(
#             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-testid='product-list-item']")),
#             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "[data-testid*='product'][data-testid*='item']"))
#         ))
#         time.sleep(0.8)
#         return True

#     except TimeoutException as e:
#         st.error(f"Timed out waiting for filtered Dutchie embed for row {row_index}. Error: {e}")
#         try:
#             driver.switch_to.default_content()
#             frames = driver.find_elements(By.TAG_NAME, "iframe")
#             st.write("Iframes on Terrabis page:", [f.get_attribute("src") for f in frames])
#             st.image(driver.get_screenshot_as_png(), caption="Terrabis page at timeout", use_container_width=True)
#         except Exception:
#             pass
#         return False


# def open_dutchie_menu(driver, wait, timeout=60):
#     """
#     Robustly enter the Dutchie menu.
#     Strategy:
#       1. Try to find the Dutchie iframe and switch to it when ready.
#       2. If switching is flaky, navigate directly to iframe src (best in headless).
#       3. As we wait, keep closing the age gate and scrolling to trigger lazy load.
#     Returns: "frame" if switched into iframe, "direct" if navigated to src.
#     Raises: TimeoutException on failure.
#     """
#     import time
#     end = time.time() + timeout
#     last_err = None

#     # helpful: try to clear overlays repeatedly while we wait
#     def _nudge_page():
#         try:
#             handle_age_verification_popup(driver, wait)
#         except Exception:
#             pass
#         try:
#             driver.execute_script("window.scrollBy(0, 800);")
#         except Exception:
#             pass
#         time.sleep(0.5)

#     while time.time() < end:
#         _nudge_page()

#         # 1) try to find the iframe
#         iframes = driver.find_elements(By.CSS_SELECTOR, "iframe[id^='dutchie--embed'], iframe[src*='dutchie']")
#         if iframes:
#             iframe = iframes[0]
#             try:
#                 # if src is set and not about:blank, directly navigate (more reliable)
#                 src = iframe.get_attribute("src")
#                 if src and "about:blank" not in src:
#                     try:
#                         driver.switch_to.default_content()
#                     except Exception:
#                         pass
#                     driver.get(src)
#                     return "direct"

#                 # otherwise, try switching into the frame and let it render
#                 wait.until(EC.frame_to_be_available_and_switch_to_it(iframe))
#                 return "frame"
#             except Exception as e:
#                 last_err = e

#         time.sleep(0.5)

#     # dump a little debug info to logs before failing
#     try:
#         frames_info = driver.execute_script(
#             "return Array.from(document.querySelectorAll('iframe'))"
#             ".map(f=>({id:f.id, src:f.src}));"
#         )
#         print("IFRAME DEBUG:", frames_info)
#     except Exception:
#         pass

#     from selenium.common.exceptions import TimeoutException
#     raise TimeoutException(f"Dutchie iframe/src not ready. Last error: {last_err}")


# TOKEN_RE = re.compile(r"""
#     \d+(?:\.\d+)?            # integer or decimal, e.g. 3 or 3.5
#     (?:\s*(?:g|mg|oz))?      # optional unit, allows a space: "3.5g" or "3.5 g"
#   | [A-Za-z]+                # or plain words
# """, re.VERBOSE | re.IGNORECASE)

# # common words to ignore
# STOPWORDS = {
#     'a','an','and','at','by','for','in','of','on','or','the','to','with', 'sample', 'hybrid', 'indica', 'sativa', 'pre', 'pod', 'popcorn', 'shake', 'pills'
# }

# # Define a list of common flavors
# FLAVOR_LIST = [
#     'apple', 'banana', 'berry', 'raspberry', 'blueberry', 'bubblegum', 'cherry', 'chocolate',
#     'citrus', 'cinnamon', 'coffee', 'cookies', 'cream', 'diesel', 'fruit', 'grape', 'lemon',
#     'lime', 'mango', 'mint', 'orange', 'peach', 'pineapple', 'lemonade', 'sour', 'strawberry', 
#     'tropical', 'vanilla', 'watermelon', 'zesty', 'sweet', 'peppermint', 'spearmint', 
#     'grapefruit', 'guava', 'spicy', 'woody', 'floral', 'gelato', 'gsc', 'haze', 
#     'zkittlez', 'runtz', 'mac', 'purp', 'gg4', 'gmo', 'shake', 'popcorn', 'mimosa'
# ]

# # Category Mapping between Excel Categories and Website Categories
# category_mapping = {
#     'BEVERAGE': 'Edibles',         # BEVERAGE to Edible on the website
#     'EDIBLE': 'Edibles',           # EDIBLE to Edible on the website
#     'PILL': 'Edibles',             # PILL to Edible on the website
#     'FLOWER': 'Flower',            # FLOWER to Flower on the website
#     'CARTRIDGE': 'Vaporizers',     # CARTRIDGE to Vaporizers on the website
#     'EXTRACT': 'Concentrates',     # EXTRACT to Concentrates on the website
#     'TOPICAL': 'Topicals',         # TOPICAL to Topicals on the website
#     'PREROLL': 'Pre-Rolls',        # PREROLL to Pre-Rolls on the website
#     'TINCTURE': 'Tinctures',       # TINCTURE to Tinctures on the website
#     'CBD': 'Tinctures',            # CBD to Tinctures on the website
#     'MERCH': 'Apparel'             # MERCH to Apparel on the website
#     # Add more mappings as needed...
# }

# # Categories on the site that have *no* weight filter
# no_weight_categories = ['Edibles', 'Topicals', 'Accessories', 'Apparel']

# # Global buffer to hold the Excel file in memory
# excel_buffer = None

# def save_updated_excel_to_memory(uploaded_file):
#     """
#     Loads the uploaded Excel file into an in-memory BytesIO buffer.
#     This buffer will be used and updated throughout the scraping process.
#     """
#     global excel_buffer
#     excel_buffer = io.BytesIO(uploaded_file.getvalue())
#     print("Excel file loaded into memory for updates.")

# def save_data_to_file(row_index, discounted_price, original_price, product_thc, product_url):
#     """
#     Updates a specific row in the in-memory Excel workbook with scraped data.
#     """
#     global excel_buffer
#     if excel_buffer is None:
#         st.error("Error: Excel buffer not initialized. Please upload a file first.")
#         return

#     try:
#         # Load the workbook from the in-memory BytesIO object
#         excel_buffer.seek(0) # Go to the beginning of the buffer
#         wb = load_workbook(excel_buffer)
        
#         # Select the specific sheet named "Pricing Research"
#         sheet_name = "Pricing Research"
#         if sheet_name not in wb.sheetnames:
#             st.warning(f"Warning: Sheet '{sheet_name}' not found. Creating it.")
#             ws = wb.create_sheet(sheet_name)
#         else:
#             ws = wb[sheet_name]

#         # openpyxl uses 1-based indexing for rows and columns
#         # The row_index from pandas is 0-based, so add 2 for Excel (1 for header, 1 for 0-base to 1-base)
#         excel_row = row_index + 2

#         # Define the target columns (AS, AT, AU, AV)
#         # AY is column 51, AZ is 52, BA is 53, BB is 54
#         col_original_price = get_column_letter(51)   # AY
#         col_discounted_price = get_column_letter(52) # AZ
#         col_thc_content = get_column_letter(53)      # BA
#         col_product_url = get_column_letter(54)      # BB

#         # Handle multiple values by joining them with commas
#         ws[f"{col_discounted_price}{excel_row}"] = ", ".join(map(str, discounted_price)) if isinstance(discounted_price, list) else discounted_price
#         ws[f"{col_original_price}{excel_row}"] = ", ".join(map(str, original_price)) if isinstance(original_price, list) else original_price
#         ws[f"{col_thc_content}{excel_row}"] = ", ".join(map(str, product_thc)) if isinstance(product_thc, list) else product_thc
#         ws[f"{col_product_url}{excel_row}"] = ", ".join(map(str, product_url)) if isinstance(product_url, list) else product_url

#         # Save the modified workbook back to the BytesIO buffer
#         new_buffer = io.BytesIO()
#         wb.save(new_buffer)

#         # Reset pointer so future reads start from the beginning
#         new_buffer.seek(0)

#         # Update the global buffer
#         excel_buffer = new_buffer

#         print(f"Row {excel_row} updated in memory for product at index {row_index}.")

#     except Exception as e:
#         st.error(f"Error saving data to Excel for row {row_index}: {e}")

# def get_driver(headful: bool = False, proxy: str | None = None):
#     import undetected_chromedriver as uc
#     from selenium.webdriver.support.ui import WebDriverWait

#     chrome_bin = _find_chrome_binary()
#     if not chrome_bin:
#         raise FileNotFoundError("No Chrome/Chromium binary found. Install it or set UC_CHROME_BINARY.")

#     os.environ["UC_CHROME_BINARY"] = chrome_bin
#     major = _chrome_major(chrome_bin)

#     options = uc.ChromeOptions()

#     # --- headful/headless toggles ---
#     if headful:
#         options.add_argument("--start-maximized")
#         try:
#             options.add_experimental_option("detach", True)
#         except Exception:
#             pass
#         options.add_argument("--no-sandbox")
#         options.add_argument("--disable-dev-shm-usage")
#     else:
#         options.add_argument("--headless=new")
#         options.add_argument("--no-sandbox")
#         options.add_argument("--disable-dev-shm-usage")
#         options.add_argument("--disable-gpu")
#         options.add_argument("--hide-scrollbars")

#     # --- prefs & page-load strategy ---
#     options.add_experimental_option("prefs", {
#         "profile.default_content_setting_values.geolocation": 1
#     })
#     try:
#         options.page_load_strategy = "eager"
#     except Exception:
#         pass

#     # --- hardening/realism ---
#     options.add_argument("--window-size=1920,1080")
#     options.add_argument("--force-device-scale-factor=1")
#     options.add_argument("--disable-features=IsolateOrigins,site-per-process")
#     options.add_argument("--disable-blink-features=AutomationControlled")
#     options.add_argument("--enable-precise-memory-info")
#     options.add_argument("--disable-extensions")
#     options.add_argument("--disable-infobars")
#     options.add_argument("--lang=en-US")
#     options.add_argument(
#         "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
#         "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
#     )

#     # --- optional proxy (server-side) ---
#     # Examples: proxy="http://host:port", proxy="socks5://host:1080"
#     if proxy:
#         options.add_argument(f"--proxy-server={proxy}")

#     driver = uc.Chrome(
#         options=options,
#         version_main=major,
#         patcher_force_close=True
#     )
#     wait = WebDriverWait(driver, 20)

#     # Extra realism via CDP (timezone/locale/UA hints + geolocation)
#     try:
#         driver.execute_cdp_cmd("Emulation.setTimezoneOverride", {"timezoneId": "America/Chicago"})
#         driver.execute_cdp_cmd("Emulation.setLocaleOverride", {"locale": "en-US"})
#         driver.execute_cdp_cmd("Network.setUserAgentOverride", {
#             "userAgent": driver.execute_script("return navigator.userAgent"),
#             "acceptLanguage": "en-US,en;q=0.9",
#             "platform": "Windows"
#         })
#         for origin in ("https://dutchie.com", "https://www.dutchie.com"):
#             driver.execute_cdp_cmd("Browser.grantPermissions", {
#                 "origin": origin,
#                 "permissions": ["geolocation"]
#             })
#         driver.execute_cdp_cmd("Emulation.setGeolocationOverride", {
#             "latitude": 38.4142, "longitude": -88.0039, "accuracy": 50
#         })
#     except Exception:
#         pass

#     return driver, wait

    
# # ----------------------------^
# # def get_driver(headful: bool = False):
# #     import undetected_chromedriver as uc
# #     from selenium.webdriver.support.ui import WebDriverWait

# #     chrome_bin = _find_chrome_binary()
# #     if not chrome_bin:
# #         raise FileNotFoundError(
# #             "No Chrome/Chromium binary found. Install it or set UC_CHROME_BINARY."
# #         )

# #     os.environ["UC_CHROME_BINARY"] = chrome_bin
# #     major = _chrome_major(chrome_bin)

# #     options = uc.ChromeOptions()

# #     # ✅ Let geolocation prompts auto-allow (some menus filter by location)
# #     options.add_experimental_option("prefs", {
# #         "profile.default_content_setting_values.geolocation": 1
# #     })

# #     # ✅ Faster DOM “ready” (optional)
# #     try:
# #         options.page_load_strategy = "eager"
# #     except Exception:
# #         pass

# #     if headful:
# #         # headed
# #         options.add_argument("--start-maximized")
# #         # keep the window open after the script finishes (nice for local debug)
# #         try:
# #             options.add_experimental_option("detach", True)
# #         except Exception:
# #             pass
# #         # On Linux you often still need these two for containerized runs
# #         options.add_argument("--no-sandbox")
# #         options.add_argument("--disable-dev-shm-usage")
# #     else:
# #         # headless
# #         options.add_argument("--headless=new")
# #         options.add_argument("--no-sandbox")
# #         options.add_argument("--disable-dev-shm-usage")
# #         options.add_argument("--disable-gpu")
# #         options.add_argument("--hide-scrollbars")

# #     options.add_argument("--window-size=1920,1080")
# #     options.add_argument("--force-device-scale-factor=1")
# #     options.add_argument("--disable-features=IsolateOrigins,site-per-process")
# #     options.add_argument("--disable-blink-features=AutomationControlled")
# #     options.add_argument("--enable-precise-memory-info")
# #     options.add_argument("--disable-extensions")
# #     options.add_argument("--disable-infobars")
# #     options.add_argument(
# #         "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
# #         "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
# #     )

# #     driver = uc.Chrome(
# #         options=options,
# #         version_main=major,
# #         patcher_force_close=True
# #     )
# #     wait = WebDriverWait(driver, 20)
# #     return driver, wait
# # ----------------------------^
    
# # def get_driver():
# #     """
# #     Headless-safe UC Chrome for Streamlit Cloud / Linux.
# #     """
# #     import os
# #     import undetected_chromedriver as uc
# #     from selenium.webdriver.support.ui import WebDriverWait

# #     # On Streamlit Cloud we install chromium via packages.txt
# #     # You can also set this in your app before calling get_driver()
# #     os.environ.setdefault("UC_CHROME_BINARY", "/usr/bin/chromium")

# #     options = uc.ChromeOptions()
# #     # — Headless & sandboxing —
# #     options.add_argument("--headless=new")
# #     options.add_argument("--no-sandbox")
# #     options.add_argument("--disable-dev-shm-usage")
# #     options.add_argument("--disable-gpu")
# #     options.add_argument("--window-size=1920,1080")
# #     options.add_argument("--force-device-scale-factor=1")
# #     options.add_argument("--disable-features=IsolateOrigins,site-per-process")
# #     options.add_argument("--disable-blink-features=AutomationControlled")

# #     # — Make layout deterministic in headless —
# #     options.add_argument("--hide-scrollbars")
# #     options.add_argument("--enable-precise-memory-info")
# #     options.add_argument("--disable-extensions")
# #     options.add_argument("--disable-infobars")

# #     # — Stable UA (helps sites that alter layout for bots) —
# #     options.add_argument(
# #         "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
# #         "AppleWebKit/537.36 (KHTML, like Gecko) "
# #         "Chrome/120.0.0.0 Safari/537.36"
# #     )


    
#     def _chromium_major():
#         # Works on Streamlit Cloud when 'chromium' is installed via packages.txt
#         try:
#             out = subprocess.check_output(["/usr/bin/chromium", "--version"]).decode()
#         except Exception:
#             # Fallback path name some images use
#             out = subprocess.check_output(["chromium", "--version"]).decode()
#         # e.g., "Chromium 120.0.6099.224"
#         m = re.search(r"(\d+)\.", out)
#         return int(m.group(1)) if m else 120
    
#     os.environ.setdefault("UC_CHROME_BINARY", "/usr/bin/chromium")
#     major = _chromium_major()
    
#     driver = uc.Chrome(
#         options=options,
#         version_main=major,          # ← tell UC which Chrome we're using
#         patcher_force_close=True     # ← safer cleanup
#     )
#     wait = WebDriverWait(driver, 20)
#     return driver, wait


# def clean_thc_value(thc_string):
#     """
#     Extracts only the value part from a THC string like 'THC: 100 mg'.
#     Returns 'N/A' if the input is 'N/A' or doesn't match the expected format.
#     """
#     if thc_string == "N/A":
#         return "N/A"

#     match = re.search(r"THC:\s*(.*)", thc_string, re.IGNORECASE)
#     if match:
#         return match.group(1).strip()
#     return thc_string # Return original if no match (e.g., if it's already just the value)

# def handle_age_verification_popup(driver, wait):
#     """
#     Close the age verification / popup reliably in headless.
#     """
#     st.info("Waiting for pop-up to appear and attempting to close it...")
#     time.sleep(2)
#     try:
#         btn = wait_present(driver, (By.CSS_SELECTOR, "a.pum-close.elementor-element-ebd2f15"), timeout=15)
#         stable_click(driver, btn)
#         time.sleep(1.0)
#         print("Pop-up closed successfully!")
#     except Exception as e:
#         print(f"Age-gate close not found or already closed: {e}")

# # def handle_age_verification_popup(driver, wait):
# #     """
# #     Handles the age verification pop-up on the Terrabis Grayville Dispensary website.
# #     Clicks the "yes, I'm 21 or older" button if the pop-up appears.
# #     """
# #     st.info("Waiting for pop-up to appear and attempting to close it...")
# #     time.sleep(4)  # Initial wait to allow the page to load
# #     try:
# #         # Wait for the "yes, I'm 21 or older" button to be clickable
# #         age_verification_button = wait.until(
# #             EC.element_to_be_clickable((By.CSS_SELECTOR, "a.pum-close.elementor-element-ebd2f15"))
# #         )
# #         age_verification_button.click()  # Click the age verification button
# #         print("Pop-up closed successfully!")
# #         time.sleep(3)  # Increased delay after closing pop-up to allow page to settle
# #     except Exception as popup_e:
# #         print(f"Could not find or click the pop-up close button. It might not have appeared or its locator changed. Error: {popup_e}")
# #         st.info("Continuing without closing pop-up. You might need to close it manually.")

# # Define direct links for specific categories
# direct_category_links = {
#     'Merch': 'https://terrabis.co/order-online/grayville/?dtche%5Bsortby%5D=relevance&dtche%5Bcategory%5D=apparel',
#     'Accessories': 'https://terrabis.co/order-online/grayville/?dtche%5Bsortby%5D=relevance&dtche%5Bcategory%5D=accessories'
# }

# def scrape_category(category, driver):
#     website_url = 'https://terrabis.co/illinois/grayville/'
    
#     # Check if a direct link is available for the current category
#     # Use .capitalize() to match how categories are usually handled or mapped
#     normalized_category = category.capitalize() 

#     if normalized_category in direct_category_links:
#         direct_link = direct_category_links[normalized_category]
#         st.info(f"Directly navigating to URL for category '{category}': {direct_link}")
#         print(f"Directly navigating to URL for category '{category}': {direct_link}")
#         driver.get(direct_link)
#         wait = WebDriverWait(driver, 20)
#         handle_age_verification_popup(driver, wait)
#         print(f"Successfully navigated to direct link for category '{category}'.")
#         st.success(f"Category '{category}' selected via direct link.")
#         time.sleep(3) # Give some time for the new category page to load
#         return True

#     # If no direct link, proceed with carousel navigation logic
#     driver.get(website_url)

#     wait = WebDriverWait(driver, 20) # Increased wait time for initial load
#     handle_age_verification_popup(driver, wait)

#     st.info(f"Attempting to select category '{category}' on the website using carousel...")

#     website_category_name = category_mapping.get(category.upper(), category.capitalize())
#     category_found = False
#     max_next_clicks = 3  # Increased limit for next clicks
#     clicked_next_count = 0

#     while not category_found and clicked_next_count <= max_next_clicks:
#         try:
#             # Construct XPath to find the <a> tag that contains an <h2> with the capitalized category name
#             category_xpath = f"//div[@class='category-slick-content']//h2[text()='{website_category_name}']/ancestor::a"

#             # Try to find the category link in the current view
#             category_link = wait_visible(driver, (By.XPATH, category_xpath))
#             if not stable_click(driver, category_link):
#                 st.warning("Category click fallback retry...")
#                 time.sleep(0.6)
#                 category_link = wait_visible(driver, (By.XPATH, category_xpath))
#                 stable_click(driver, category_link)
#             print(f"Successfully selected category '{category}' on the website.")
#             st.success(f"Category '{category}' found and selected.")
#             time.sleep(3) # Give some time for the new category page to load
#             category_found = True
#             return True

#         except (NoSuchElementException, TimeoutException):
#             # Category not found in the current view. Try clicking 'Next'.
#             print(f"Category '{category}' not found in current view. Looking for 'Next' button...")
#             st.info(f"Category '{category}' not found, trying to click 'Next' to reveal more categories.")
            
#             next_button_click_attempts = 3 # Max attempts to click the 'Next' button
#             for attempt in range(next_button_click_attempts):
#                 try:
#                     # Locate the "Next" button
#                     next_button = wait_visible(driver, (By.CSS_SELECTOR, "button.slick-next.slick-arrow[aria-label='Next'][type='button']"))
#                     # Check disabled state
#                     if next_button.get_attribute("aria-disabled") == "true":
#                         print("Reached the end of categories, 'Next' button is disabled.")
#                         st.warning("Reached the end of categories, desired category not found.")
#                         category_found = False
#                         return False
                    
#                     # Ensure in view
#                     driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_button)
#                     time.sleep(0.2)
#                     # Track carousel movement
#                     slick_track = wait_present(driver, (By.CSS_SELECTOR, ".slick-track"))
#                     initial_transform = slick_track.get_attribute("style")
                    
#                     stable_click(driver, next_button)
#                     WebDriverWait(driver, 10).until(
#                         lambda d: d.find_element(By.CSS_SELECTOR, ".slick-track").get_attribute("style") != initial_transform
#                     )
#                     time.sleep(0.6)
                    
#                     clicked_next_count += 1  # <-- add this line so the outer loop progresses
#                     # next_button = wait.until(
#                     #     EC.element_to_be_clickable((By.CSS_SELECTOR, "button.slick-next.slick-arrow[aria-label='Next'][type='button']"))
#                     # )
                    
#                     # # Check if the button is disabled
#                     # if next_button.get_attribute("aria-disabled") == "true":
#                     #     print("Reached the end of categories, 'Next' button is disabled.")
#                     #     st.warning("Reached the end of categories, desired category not found.")
#                     #     category_found = False # Ensure loop terminates
#                     #     return False # Exit early if no more categories
                    
#                     # # Scroll the button into view to ensure it's actionable
#                     # driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
#                     # time.sleep(1) # Small pause after scrolling
                    
#                     # # Get the initial transform style of the slick-track before clicking
#                     # slick_track = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".slick-track")))
#                     # initial_transform = slick_track.get_attribute("style")
                    
#                     # # Attempt JavaScript click
#                     # driver.execute_script("arguments[0].click();", next_button)
                    
#                     # clicked_next_count += 1
#                     # print(f"Clicked 'Next' button. Attempt {clicked_next_count}/{max_next_clicks}. Click retry: {attempt + 1}/{next_button_click_attempts}")
                    
#                     # # IMPORTANT: Wait for the 'transform' style of the slick-track to change
#                     # # This indicates the carousel has actually moved.
#                     # wait.until(lambda d: d.find_element(By.CSS_SELECTOR, ".slick-track").get_attribute("style") != initial_transform)
                    
#                     time.sleep(2) # Additional brief pause after content update and carousel movement
#                     break # Break from the retry loop if click was successful and content updated
                
#                 except (ElementClickInterceptedException, StaleElementReferenceException) as e:
#                     print(f"Click interception/stale element on 'Next' button. Retrying (Attempt {attempt + 1}/{next_button_click_attempts}). Error: {e}")
#                     st.warning(f"Next button click intercepted or stale. Retrying ({attempt + 1}/{next_button_click_attempts})...")
#                     time.sleep(2) # Wait before retrying the click
#                     if attempt == next_button_click_attempts - 1:
#                         print("Max retry attempts for 'Next' button reached. Carousel did not advance.")
#                         st.error("Could not click 'Next' button or carousel did not advance after multiple attempts.")
#                         return False # Give up if retries fail
#                 except (NoSuchElementException, TimeoutException) as next_button_e:
#                     print(f"No 'Next' button found or it's not clickable after previous attempts. Error: {next_button_e}")
#                     st.error("No more 'Next' categories or button not found. Category not available.")
#                     return False # Exit if Next button is truly not there

#             else: # This else block executes if the inner for loop completes without a 'break'
#                 # This means the 'Next' button click attempts failed without an explicit exception
#                 # indicating the end of the line (e.g., if it never became clickable)
#                 print("Failed to click 'Next' button after all retries or carousel did not advance.")
#                 st.error("Failed to click 'Next' button after all retries or carousel did not advance.")
#                 return False # Exit function if click loop fails

#         except Exception as category_e:
#             print(f"An unexpected error occurred while trying to select category '{category}'. Error: {category_e}")
#             st.error(f"An unexpected error occurred: {category_e}")
#             return False

#     if not category_found:
#         print(f"Failed to select category '{category}' after trying to click 'Next' {clicked_next_count} times.")
#         st.error(f"Category '{category}' not available on website after checking all visible categories.")
#         return False

#     return True

# # Brand mapping between Excel and Website names
# brand_mapping = {       
#     'FLORACAL': 'FloraCal Farms',
#     'NATURE\'S GRACE & WELLNESS': 'Nature\'s Grace and Wellness',
#     'WANA GUMMIES': 'Wana',
#     'WONDER WELLNESS': 'Wonder',
#     'UPNORTH HUMBOLT': 'UpNorth',
#     'LULA': 'Lula\'s',
#     'JOOS': 'Joos Vapes',
#     'MIDWEEK FRIDAY': 'Mid Week Friday'    
#     # Continue adding brand mappings as needed.
# }

# # Categories on the site that have no brand filter
# no_brand_categories = ['Apparel']

# def scrape_brand(brand, driver):
#     """
#     Selects a brand using a headless-safe approach:
#     1) Expand "Brands" section
#     2) Try React-safe typing into the search box (dispatch events)
#     3) Fallback: click 'Show more' and scan/scroll labels
#     """
#     brand_name_on_website = brand_mapping.get(brand, brand)
#     target_norm = " ".join(brand_name_on_website.lower().split())

#     # If category doesn't use brand filters, do the simple direct scan you had
#     if getattr(driver, "current_category", None) in ["TOPICAL", "ACCESSORIES"]:
#         labels = driver.find_elements(By.XPATH, "//label")
#         for lbl in labels:
#             label_text = " ".join(lbl.text.split()).lower()
#             if target_norm in label_text:
#                 cb = driver.find_element(By.CSS_SELECTOR, f"input[id='{lbl.get_attribute('for')}']")
#                 if not cb.is_selected():
#                     driver.execute_script("arguments[0].scrollIntoView({block:'center'});", lbl)
#                     time.sleep(0.2)
#                     stable_click(driver, lbl)
#                     print(f"✔ Selected brand (direct): {lbl.text}")
#                 else:
#                     print(f"ℹ️ Already selected (direct): {lbl.text}")
#                 return True
#         st.error(f"⚠️ Brand not found (direct): {brand_name_on_website}")
#         return False

#     # Expand "Brands" filter
#     try:
#         brand_section_button = driver.find_element(By.XPATH, "//button[contains(., 'Brands')]")
#         driver.execute_script("arguments[0].scrollIntoView({block:'center'});", brand_section_button)
#         time.sleep(0.5)
#         if brand_section_button.get_attribute("aria-expanded") == "false":
#             stable_click(driver, brand_section_button)
#             time.sleep(0.8)
#         # Keep the section in view
#         driver.execute_script("arguments[0].scrollIntoView({block:'center'});", brand_section_button)
#         time.sleep(0.3)
#     except Exception as e:
#         print(f"Could not expand 'Brands' section: {e}")

#     # Try search input first (React-safe)
#     try:
#         # Avoid hashed classes; use placeholder/role based selectors
#         search_input = WebDriverWait(driver, 6).until(EC.presence_of_element_located((
#             By.CSS_SELECTOR,
#             "input[placeholder*='earch'][type='text'], input[placeholder*='Brand'], input[type='search']"
#         )))
#         # Focus and clear via JS, then React-safe set
#         driver.execute_script("arguments[0].focus();", search_input)
#         driver.execute_script("arguments[0].value='';", search_input)
#         type_react_input(driver, search_input, brand_name_on_website)
#         time.sleep(1.0)  # let debounce/filtering run

#         # Wait for matching label to appear
#         label = WebDriverWait(driver, 6).until(EC.presence_of_element_located((
#             By.XPATH,
#             f"//label[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{target_norm}')]"
#         )))
#         driver.execute_script("arguments[0].scrollIntoView({block:'center'});", label)
#         time.sleep(0.2)
#         stable_click(driver, label)

#         # Ensure checkbox toggled
#         try:
#             for_attr = label.get_attribute("for")
#             cb = driver.find_element(By.CSS_SELECTOR, f"input[id='{for_attr}']")
#             if not cb.is_selected():
#                 stable_click(driver, label)
#         except Exception:
#             pass

#         print(f"✔ Selected brand via search: {brand_name_on_website}")
#         time.sleep(1.0)
#         return True

#     except Exception as e:
#         print(f"Search-based selection failed (falling back): {e}")

#     # Fallbacks when search is unreliable in headless
#     click_show_more_until_exhausted(driver)
#     try:
#         ok = scroll_filter_panel_to_find_label(driver, brand_name_on_website, max_scrolls=40)
#         if ok:
#             print(f"✔ Selected brand via list scan: {brand_name_on_website}")
#             time.sleep(1.0)
#             return True
#         else:
#             st.error(f"⚠️ Brand not found: {brand_name_on_website}")
#             return False
#     except Exception as e:
#         print(f"Label-scan fallback failed: {e}")
#         st.error(f"⚠️ Could not select brand '{brand_name_on_website}'.")
#         return False


# def normalize_weight(weight, has_unit=False):
#     """
#     Normalize the weight format to match the website format.
#     Converts weight from GRAMS to g, and MILLIGRAMS to mg.
#     """
#     weight = str(weight).lower().strip()
    
#     # Replace 'grams' with 'g' and 'milligrams' with 'mg'
#     weight = weight.replace("grams", "g").replace("milligrams", "mg").replace("millig", "mg").replace(" ", "")
    
#     # Handle cases like ".7g" instead of "0.7g"
#     if weight.startswith("."):
#         weight = "0" + weight  # Converts ".7g" to "0.7g"
    
#     return weight

# def grams_to_ounces(grams):
#     """
#     Converts grams to ounces and returns the fractional representation (e.g., 3.5g -> 1/8oz).
#     """
#     ounces = grams * 0.03527396 # Convert grams to ounces
    
#     # Convert ounces to a fraction (e.g., 0.125oz -> 1/8oz)
#     if ounces == 0.125:
#         return '1/8oz'
#     elif ounces == 0.25:
#         return '1/4oz'
#     elif ounces == 0.5:
#         return '1/2oz'
#     elif ounces == 1:
#         return '1oz'
#     elif ounces == 2:
#         return '2oz'
#     else:
#         # For complex cases, round to a practical number and convert to the closest fraction
#         rounded_ounces = round(ounces, 2)
#         # Now we can handle small fractions that might not directly fit the exact numbers above
#         if abs(rounded_ounces - 0.125) < 0.02:
#             return '1/8oz'
#         elif abs(rounded_ounces - 0.25) < 0.02:
#             return '1/4oz'
#         elif abs(rounded_ounces - 0.5) < 0.02:
#             return '1/2oz'
#         else:
#             # For any other case (like 0.7 or 0.12oz), return the decimal representation
#             return f"{rounded_ounces}oz"

# def scrape_weight(weight, driver):
#     """
#     Selects the weight filter from the weight options.
#     This function clicks the weight option based on the provided weight value.
#     """
#     # Normalize weight (e.g., from "0.75 GRAMS" or "1 GRAMS")
#     weight_norm = normalize_weight(weight)
#     # also try dropping a leading zero so "0.75g" → ".75g"
#     variants = [weight_norm]
#     if weight_norm.startswith("0."):
#         variants.append(weight_norm[1:])

#     try:
#         # Find all weight filter links
#         weight_links = driver.find_elements(By.CSS_SELECTOR, "a.weight__Anchor-sc-10b36p8-0.geHygR")

#         # Loop through each weight option
#         for link in weight_links:
#             link_text = link.text.strip().lower()

#             # assume pure numbers are grams (e.g. "28" -> "28g")
#             if link_text.replace('.', '', 1).isdigit():
#                 link_text = link_text + 'g'

#             # try each of our variants (with and without leading zero)
#             for v in variants:
#                 if v in link_text:
#                     stable_click(driver, link)
#                     print(f"✔ Selected weight: {link_text}")
#                     return True

#         # If no matching weight found in grams, convert to ounces
#         print(f"⚠️ Weight '{weight}' not found in grams. Trying to convert to ounces...")

#         # Convert weight to ounces
#         weight_in_ounces = grams_to_ounces(float(weight.replace('g', '').strip()))
#         print(f"Converted weight: {weight_in_ounces}")

#         # Search for the ounce weight
#         for link in weight_links:
#             link_text = link.text.strip().lower()
#             if weight_in_ounces in link_text:
#                 stable_click(driver, link)
#                 print(f"✔ Selected weight in ounces: {link_text}")
#                 return True

#         # If still not found
#         st.error(f"⚠️ Weight '{weight_in_ounces}' not found.")
#         print(f"⚠️ Weight '{weight_in_ounces}' not found.")
#         return False

#     except Exception as e:
#         print(f"⚠️ Could not select weight '{weight}'. Error: {e}")
#         return False

# QUANTITY_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(pack|pk|ct|capsules|capsule|count|ea|unit|qty)", re.IGNORECASE)

# def extract_and_normalize_quantity(text):
#     """
#     Extracts numerical quantity and normalizes the unit.
#     Returns (quantity_number, normalized_unit_type) or (None, None).
#     Normalized unit types: 'mg', 'ml', 'count'
#     """
#     match = QUANTITY_RE.search(text)
#     if match:
#         number = float(match.group(1))
#         unit = match.group(2).lower()
#         if unit in ['pack', 'pk', 'ct', 'capsules', 'capsule', 'count', 'ea', 'unit', 'qty']:
#             return number, 'count'
#     return None, None

# RATIO_RE = re.compile(
#     r"""
#     # Optional opening parenthesis or bracket, followed by optional whitespace
#     (?:[\(\[]?\s*)?
#     (?:
#         # Pattern A: Cannabinoid(s) first, then numerical ratio (e.g., THC:CBD 1:2, THC/CBD/CBG 1:1:1)
#         (?:THC|CBD|CBG|CBN|CBC)(?:[\s\/:]?(?:THC|CBD|CBG|CBN|CBC))* \s* \d+[:\/]\d+(?:[:\/]\d+)*
#         |
#         # Pattern B: Numerical ratio first, then Cannabinoid(s) (e.g., 1:1 THC:CBG, 8:1 CBD/THC)
#         \d+[:\/]\d+(?:[:\/]\d+)* \s* (?:THC|CBD|CBG|CBN|CBC)(?:[\s\/:]?(?:THC|CBD|CBG|CBN|CBC))*
#     )
#     # Optional whitespace followed by optional closing parenthesis or bracket
#     (?:\s*[\)\]]?)?
#     """,
#     re.VERBOSE | re.IGNORECASE
# )

# def extract_flavors(text, flavor_list):
#     """
#     Extracts flavors from a given text based on a predefined list of flavors.
#     Returns a list of unique flavors found.
#     """
#     found_flavors = set()
#     text_lower = text.lower()
    
#     # Sort flavors by length descending to match longer phrases first (e.g., "blue raspberry" before "blue")
#     sorted_flavors = sorted(flavor_list, key=len, reverse=True)

#     for flavor in sorted_flavors:
#         # Use word boundaries to avoid partial matches (e.g., 'grape' in 'grapefruit')
#         if re.search(r'\b' + re.escape(flavor) + r'\b', text_lower):
#             found_flavors.add(flavor)
#     return list(found_flavors)

# def word_match_score(a, b):
#     """
#     Returns fraction of words in a that also appear in b.
#     """
#     wa = re.findall(r"\w+", a.lower())
#     wb = set(re.findall(r"\w+", b.lower()))
#     if not wa:
#         return 0
#     matches = sum(1 for w in wa if w in wb)
#     return matches / len(wa)

# # Custom CSS to style the app
# st.markdown("""
#     <style>
#         /* Center the title */
#         .title {
#             text-align: center;
#             color: #000000;
#             font-size: 32px;
#         }

#         /* Style the main page background */
#         .stApp {
#             background-color: #F4F7FC; /* Light Gray-Blue */
#         }
        
#         .title-container h1 {
#             color: white !important; /* Force White Text */
#             font-size: 32px;
#             font-weight: bold;
#         }

#         /* Style the header background */
#         .stMarkdown h1 {
#             background-color: #6a2af1; /* Dark Blue */
#             color: #fff;
#             padding: 15px;
#             border-radius: 8px;
#             text-align: center;
#         }

#         /* Style the sidebar */
#         .css-1d391kg {
#             background-color: #EAECEF; /* Light Gray */
#             padding: 20px;
#             border-radius: 10px;
#         }

#         /* Buttons */
#         .stButton>button {
#             background-color: #1E3A8A; /* Dark Blue */
#             color: white;
#             border-radius: 5px;
#             padding: 10px;
#             font-weight: bold;
#             border: none;
#             transition: 0.3s;
#         }

#         .stButton>button:hover {
#             background-color: #1E40AF; /* Slightly Brighter Blue */
#         }

#         /* Style file uploader */
#         .stFileUploader {
#             background-color: white;
#             border-radius: 8px;
#             padding: 10px;
#             border: 1px solid #B0BEC5; /* Gray Border */
#         }

#         /* Style the download button */
#         .stDownloadButton>button {
#             background-color: #10B981; /* Green */
#             color: white;
#             border-radius: 5px;
#             padding: 10px;
#             font-weight: bold;
#             border: none;
#             transition: 0.3s;
#         }

#         .stDownloadButton>button:hover {
#             background-color: #059669; /* Darker Green */
#         }
#     </style>
# """, unsafe_allow_html=True)

# # Title at the top (centered)
# st.markdown("""
#     <div class="title-container">
#         <h1 class="title-text">Web Scraper for Terrabis Data</h1>
#     </div>
# """, unsafe_allow_html=True)

# # Streamlit App Interface
# st.sidebar.title('Instructions')

# # Display instructions in the sidebar
# st.sidebar.markdown("""
#     1. **Upload an Excel file** that contains product pricing data for scraping.
#     2. Click **Start Scraping** to begin the extraction process.
#     3. The scraper will extract product details and update the file.
#     4. After the scraping process, you will be able to download the updated file.
#     5. The scraper will process different product categories.
# """)

# # File Upload in Sidebar
# uploaded_file = st.sidebar.file_uploader("Upload Excel File", type=['xlsx'])

# if uploaded_file:
#     # Load the file into memory buffer *once* when uploaded
#     save_updated_excel_to_memory(uploaded_file) 
    
#     # Load the file from the BUFFER
#     excel_buffer.seek(0)
#     df = pd.read_excel(excel_buffer, sheet_name="Pricing Research")
    
#     # Extract unique categories and brands from the 'Category' and 'Brand' columns
#     categories = df['Category'].unique()
#     brands = df['Brand'].unique()

#     # Dropdown in Sidebar for Category Selection
#     selected_category = st.sidebar.selectbox("Select Category to Scrap", categories)
    
#     # Filter the data based on the selected category
#     filtered_data = df[df['Category'] == selected_category]

#     # Show the number of products in the selected category
#     num_products = len(filtered_data)
#     st.write('-------------------------------------------------------------------------')
#     st.write(f"Number of products in the '{selected_category}' category: {num_products}")

#     # Button to Start Scraping
#     if st.sidebar.button('Start Scraping'):
#         st.write("Scraping started for category:", selected_category)

#         # Initialize the driver once and pass it to both category and brand selection functions
#         driver, wait = get_driver()

#         driver.current_category = selected_category

#         # Start the category selection process
#         category_found = scrape_category(selected_category, driver)

#         if not category_found:
#             # Quit the driver and stop the script if the category wasn't found
#             driver.quit()
#             st.warning("Scraping stopped because the category was not found on the website.")
#             # Use 'st.stop()' to halt execution in a Streamlit app
#             st.stop()

#         # Store the current category URL to reload it later
#         category_url = driver.current_url
#         print("Category URL stored:", category_url)

#         # Filter brands based on the selected category and scrape
#         relevant_brands = df[df['Category'] == selected_category]['Brand'].tolist()

#         # Loop through the brands and select them
#         for row_index, row in filtered_data.iterrows():
#             # Reload the category page so all filters are cleared
#             driver.switch_to.default_content()
#             driver.get(category_url)

#             # --- START NEW CODE BLOCK: Handle IFRAME and Wait for Products ---
#             # --- START NEW CODE BLOCK: Handle IFRAME and Wait for Products ---
#             # try:
#             #     time.sleep(1.0)
            
#             #     # Find the Dutchie iframe (and capture its src for a fallback)
#             #     iframe = WebDriverWait(driver, 15).until(
#             #         EC.presence_of_element_located((
#             #             By.CSS_SELECTOR,
#             #             "iframe#dutchie--embed__iframe, iframe[id*='dutchie'], iframe[src*='dutchie.com']"
#             #         ))
#             #     )
#             #     iframe_src = iframe.get_attribute("src") or ""
#             #     WebDriverWait(driver, 10).until(lambda d: iframe.get_attribute("src") and "about:blank" not in iframe.get_attribute("src"))
            
#             #     # Try switching INTO the iframe first
#             #     try:
#             #         WebDriverWait(driver, 15).until(EC.frame_to_be_available_and_switch_to_it(iframe))
#             #         st.info(f"Switched to Dutchie iframe for row {row_index}.")
            
#             #         # DOM ready, then wait for products
#             #         WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
            
#             #         # Close cookie/banner if it appears
#             #         try:
#             #             cookie_btn = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((
#             #                 By.XPATH,
#             #                 "//button[normalize-space()='Accept' or normalize-space()='Accept all' or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept')]"
#             #             )))
#             #             stable_click(driver, cookie_btn)
#             #             time.sleep(0.4)
#             #         except Exception:
#             #             pass
            
#             #         # Wait for product tiles (several possible selectors)
#             #         WebDriverWait(driver, 25).until(EC.any_of(
#             #             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-testid='product-list-item']")),
#             #             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "[data-testid*='product'][data-testid*='item']")),
#             #             EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class,'product') and (contains(@class,'card') or contains(@class,'item'))]"))
#             #         ))
#             #         time.sleep(0.8)  # let prices/options hydrate
            
#             #     except TimeoutException:
#             #         # Fallback: open the Dutchie menu directly (avoids headless iframe issues)
#             #         driver.switch_to.default_content()
#             #         if not iframe_src:
#             #             raise  # no URL to open; let outer handler catch
            
#             #         st.info("Iframe slow/blocked; opening Dutchie menu directly.")
#             #         driver.get(iframe_src)
            
#             #         # Grant geo on dutchie.com (helps some locations)
#             #         try:
#             #             driver.execute_cdp_cmd("Browser.grantPermissions", {
#             #                 "origin": "https://dutchie.com",
#             #                 "permissions": ["geolocation"]
#             #             })
#             #             # Grayville-ish (adjust if you like)
#             #             driver.execute_cdp_cmd("Emulation.setGeolocationOverride", {
#             #                 "latitude": 38.4142, "longitude": -88.0039, "accuracy": 50
#             #             })
#             #         except Exception:
#             #             pass
            
#             #         WebDriverWait(driver, 20).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
            
#             #         # Close cookie/banner if present
#             #         try:
#             #             cookie_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
#             #                 By.XPATH,
#             #                 "//button[normalize-space()='Accept' or normalize-space()='Accept all' or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept')]"
#             #             )))
#             #             stable_click(driver, cookie_btn)
#             #             time.sleep(0.4)
#             #         except Exception:
#             #             pass
            
#             #         # Wait for product tiles on the full dutchie page
#             #         WebDriverWait(driver, 30).until(EC.any_of(
#             #             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-testid='product-list-item']")),
#             #             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "[data-testid*='product'][data-testid*='item']"))
#             #         ))
#             #         time.sleep(0.8)
            
#             # except TimeoutException as e:
#             #     st.error(f"Timed out waiting for Dutchie menu for row {row_index}. Error: {e}")
            
#             #     # Debug helpers
#             #     try:
#             #         driver.switch_to.default_content()
#             #         iframes = driver.find_elements(By.TAG_NAME, "iframe")
#             #         st.write("Iframes on page:", [f.get_attribute("src") for f in iframes])
#             #         st.image(driver.get_screenshot_as_png(), caption="Screenshot at timeout", use_container_width=True)
#             #     except Exception:
#             #         pass
            
#             #     save_data_to_file(row_index, " ", " ", " ", " ")
#             #     try:
#             #         driver.switch_to.default_content()
#             #     except Exception:
#             #         pass
#             #     continue
            
#             # except Exception as e:
#             #     st.error(f"Unexpected error while entering Dutchie for row {row_index}: {e}")
#             #     save_data_to_file(row_index, " ", " ", " ", " ")
#             #     try:
#             #         driver.switch_to.default_content()
#             #     except Exception:
#             #         pass
#             #     continue
#             # # --- END NEW CODE BLOCK ---



#             # Extract the relevant data
#             selected_category = row['Category']
#             brand = str(row['Brand'])
#             # build a set of this row’s brand tokens (collapse spaces & lowercase)
#             raw_brand = TOKEN_RE.findall(brand)
#             brand_tokens = {t.replace(" ", "").lower() for t in raw_brand}
#             website_cat      = category_mapping.get(selected_category, selected_category)
#             raw_cat          = TOKEN_RE.findall(website_cat)
#             collapsed_cat    = [t.replace(" ", "") for t in raw_cat]
#             category_tokens  = set()
#             for tok in collapsed_cat:
#                 lc = tok.lower()
#                 category_tokens.add(lc)
#                 if lc.endswith("s"):
#                     category_tokens.add(lc[:-1])
#             weight = row['Weight']  # Assuming weight is in the 'Weight' column

#             # Normalize the weight before passing it to scrape_weight
#             normalized_weight = normalize_weight(weight)

#             # Decide whether we need to select a brand
#             website_cat = category_mapping.get(selected_category, selected_category)
#             mapped_brand = brand_mapping.get(brand, brand)
            
#             if website_cat in no_brand_categories:
#                 # no brand facet on site: still open category page and switch into iframe
#                 st.info(f"Opening category via URL: {website_cat} (no brand facet).")
#                 brand_successfully_selected = open_terrabis_with_brand(
#                     driver, wait,
#                     city_slug="grayville",
#                     category_site_name=website_cat,
#                     brand_site_name=None,          # no brand param
#                     row_index=row_index
#                 )
#             else:
#                 # use URL-driven brand filter first
#                 st.info(f"Applying brand via URL: {mapped_brand} in {website_cat}")
#                 brand_successfully_selected = open_terrabis_with_brand(
#                     driver, wait,
#                     city_slug="grayville",
#                     category_site_name=website_cat,
#                     brand_site_name=mapped_brand,  # mapped name → slug
#                     row_index=row_index
#                 )
            
#                 # optional UI fallback if URL approach failed
#                 if not brand_successfully_selected:
#                     st.warning("URL brand filter failed; trying UI brand filter.")
#                     driver.switch_to.default_content()
#                     driver.get(category_url)
#                     brand_successfully_selected = scrape_brand(brand, driver)


#             # --- WEIGHT SELECTION (CONDITIONAL) ---
#             # Only proceed with weight selection if the brand was successfully selected (or skipped)
#             if brand_successfully_selected:
#                 weight_successfully_selected = False
#                 if website_cat in no_weight_categories:
#                     print(f"⏭ Skipping weight selection for category '{website_cat}' (no weight filter on site).")
#                     weight_successfully_selected = True # Consider weight "selected" if category doesn't require it
#                 else:
#                     print(f"Selecting weight: {normalized_weight}")
#                     weight_successfully_selected = scrape_weight(normalized_weight, driver)

#                 # The product scraping and matching logic will now run only if brand_successfully_selected is True
#                 # and weight_successfully_selected is True (or if they were successfully skipped).
#             else:
#                 # If brand was NOT successfully selected, skip weight selection and mark product as N/A
#                 st.error(f"⚠️ Brand '{brand}' not found")
#                 # Save empty strings for price, THC, and URL for this row
#                 save_data_to_file(row_index, " ", " ", " ", " ")
#                 continue # Skip the rest of the loop for this product and go to the next

#             # --- PRODUCT MATCHING START ---
#             try:
#                 # wait for the product tiles to appear
#                 WebDriverWait(driver, 8).until(
#                     EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-testid='product-list-item']"))
#                 )
#                 time.sleep(5)  # give a bit extra for everything to render

#                 # scrape all product names on the page
#                 product_tiles = driver.find_elements(By.CSS_SELECTOR, "div[data-testid='product-list-item']")
#                 scraped = []

#                 # grab the target from Excel
#                 target_name = row['Product Name']
#                 # Extract and normalize quantity from Excel product name
#                 excel_qty_num, excel_qty_unit = extract_and_normalize_quantity(target_name)

#                 # Identify all parts of quantity strings from the Excel product name
#                 excel_quantity_parts_to_exclude = set()
#                 for m in QUANTITY_RE.finditer(target_name):
#                     full_match_str = m.group(0).lower() # e.g., "10 pk", "100mg"
#                     # Tokenize the full match string using TOKEN_RE to get its constituent tokens
#                     # and add them to the set of parts to exclude
#                     for part_token in TOKEN_RE.findall(full_match_str):
#                         excel_quantity_parts_to_exclude.add(part_token.replace(" ", "").lower())

#                 # Identify all parts of ratio strings from the Excel product name
#                 excel_ratio_parts_to_exclude = set()
#                 for m in RATIO_RE.finditer(target_name):
#                     full_match_str = m.group(0).lower()
#                     for part_token in TOKEN_RE.findall(full_match_str):
#                         excel_ratio_parts_to_exclude.add(part_token.replace(" ", "").lower())
                
#                 # Extract flavors from Excel product name (still needed for strict matching)
#                 excel_flavors = extract_flavors(target_name, FLAVOR_LIST)
#                 excel_flavor_tokens = set(excel_flavors) # Store as set for comparison

#                 # tokenize and collapse spaces in units (so "3.5 g" → "3.5g")
#                 raw_wa       = TOKEN_RE.findall(target_name)
#                 collapsed_wa = [t.replace(" ", "") for t in raw_wa]

#                 # extract weight tokens (anything starting with a digit)
#                 excel_weight_tokens = [
#                     t for t in collapsed_wa
#                     if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', t.lower())
#                 ]

#                 # Filter Excel tokens for keywords. Flavors are *not* excluded here.
#                 excel_keyword_tokens_list = []
#                 for t in collapsed_wa:
#                     normalized_t_lower = t.lower()
#                     # Check if it's a weight token
#                     if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', normalized_t_lower):
#                         continue
#                     # Check if the token is part of an identified quantity
#                     if normalized_t_lower in excel_quantity_parts_to_exclude:
#                         continue
#                     # Check if the token is part of an identified ratio
#                     if normalized_t_lower in excel_ratio_parts_to_exclude:
#                         continue
#                     # Check if it's a brand/category/stopword token
#                     if normalized_t_lower in brand_tokens or \
#                        normalized_t_lower in STOPWORDS or \
#                        normalized_t_lower in category_tokens or \
#                        (normalized_t_lower.endswith("s") and normalized_t_lower[:-1] in category_tokens):
#                         continue
#                     # Flavors are NO LONGER EXCLUDED HERE; they will contribute to the general score.
#                     excel_keyword_tokens_list.append(normalized_t_lower)

#                 # Convert to set for efficient lookup during comparison
#                 excel_keyword_tokens_set = set(excel_keyword_tokens_list)
#                 excel_tokens_display = [t.title() for t in excel_keyword_tokens_list]

#                 st.write(f"🔎 **Product name:** {target_name}")
#                 print(f"⚖️ **Excel weight tokens:** {', '.join(excel_weight_tokens)}")
#                 print(f"📦 **Excel quantity:** {excel_qty_num} {excel_qty_unit if excel_qty_unit else 'N/A'}")
#                 print(f"🎨 **Excel flavors:** {', '.join(excel_flavors) if excel_flavors else 'N/A'}") # Display extracted flavors
#                 print(f"🔍 **Excel tokens (cleaned):** {', '.join(excel_tokens_display)}") # Now truly cleaned

#                 for tile in product_tiles:
#                     name = tile.find_element(By.CSS_SELECTOR, "div.full-card__Name-sc-11z5u35-4").text
                    
#                     # Extract URL
#                     product_url = " "
#                     try:
#                         # Attempt to find the anchor tag for the product URL
#                         url_element = tile.find_element(By.TAG_NAME, "a")
#                         product_url = url_element.get_attribute("href")
#                     except NoSuchElementException:
#                         print(f"⚠️ URL not found for product '{name}'")

#                     # Extract THC
#                     thc_content = " "
#                     try:
#                         thc_element = tile.find_element(By.CSS_SELECTOR, "div.full-card__Potency-sc-11z5u35-8 > div")
#                         raw_thc_text = thc_element.text
#                         thc_content = clean_thc_value(raw_thc_text) # Apply the cleaning function here
#                     except NoSuchElementException:
#                         pass # THC might not be present for all products

#                     # Extract Price
#                     discounted_price = " "
#                     original_price = " "
#                     try:
#                         option_tile_button = tile.find_element(By.CSS_SELECTOR, "button[data-testid='option-tile']")
                        
#                         # First, try to find the original price span, as its presence dictates the logic
#                         try:
#                             original_price_element_if_discount = option_tile_button.find_element(By.CSS_SELECTOR, "span.optionstyles__OriginalPrice-sc-vu6uvs-2")
#                             # If this element is found, it means there's a discount
#                             original_price = original_price_element_if_discount.text
#                             # The 'b' tag then holds the discounted price
#                             discounted_price_element = option_tile_button.find_element(By.TAG_NAME, "b")
#                             discounted_price = discounted_price_element.text
#                         except NoSuchElementException:
#                             # If original_price_element_if_discount is NOT found, it means no discount
#                             # In this case, the 'b' tag holds the original price
#                             single_price_element = option_tile_button.find_element(By.TAG_NAME, "b")
#                             original_price = single_price_element.text
#                             discounted_price = " " # As per new requirement
                            
#                     except NoSuchElementException:
#                         print(f"⚠️ Price information not found for product '{name}'")

#                     # Extract and normalize quantity from Site product name
#                     site_qty_num, site_qty_unit = extract_and_normalize_quantity(name)

#                     # Identify all parts of quantity strings from the Site product name
#                     site_quantity_parts_to_exclude = set()
#                     for m in QUANTITY_RE.finditer(name):
#                         full_match_str = m.group(0).lower()
#                         for part_token in TOKEN_RE.findall(full_match_str):
#                             site_quantity_parts_to_exclude.add(part_token.replace(" ", "").lower())

#                     # Identify all parts of ratio strings from the Site product name
#                     site_ratio_parts_to_exclude = set()
#                     for m in RATIO_RE.finditer(name):
#                         full_match_str = m.group(0).lower()
#                         for part_token in TOKEN_RE.findall(full_match_str):
#                             site_ratio_parts_to_exclude.add(part_token.replace(" ", "").lower())
                    
#                     # Extract flavors from Site product name
#                     site_flavors = extract_flavors(name, FLAVOR_LIST)
#                     site_flavor_tokens = set(site_flavors) # Store as set for comparison

#                     raw_wb = TOKEN_RE.findall(name)
#                     collapsed_wb = [t.replace(" ", "") for t in raw_wb]

#                     site_weight_tokens = [
#                         t for t in collapsed_wb
#                         if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', t.lower())
#                     ]

#                     # Filter Site tokens for keywords. Flavors are *not* excluded here.
#                     site_keyword_tokens_list = []
#                     for t in collapsed_wb:
#                         normalized_t_lower = t.lower()
#                         # Check if it's a weight token
#                         if re.match(r'^\d+(?:\.\d+)?(?:g|mg|oz)$', normalized_t_lower):
#                             continue
#                         # Check if it's part of an identified quantity
#                         if normalized_t_lower in site_quantity_parts_to_exclude:
#                             continue
#                         # Check if the token is part of an identified ratio
#                         if normalized_t_lower in site_ratio_parts_to_exclude:
#                             continue
#                         # Check if it's a brand/category/stopword token
#                         if normalized_t_lower in brand_tokens or \
#                            normalized_t_lower in STOPWORDS or \
#                            normalized_t_lower in category_tokens or \
#                            (normalized_t_lower.endswith("s") and normalized_t_lower[:-1] in category_tokens):
#                             continue
#                         # Flavors are NO LONGER EXCLUDED HERE; they will contribute to the general score.
#                         site_keyword_tokens_list.append(normalized_t_lower)

#                     # Add site_qty_num, site_qty_unit, and site_keyword_tokens_list to scraped tuple
#                     scraped.append((name, product_url, discounted_price, original_price, thc_content, site_weight_tokens, site_qty_num, site_qty_unit, site_flavors, site_keyword_tokens_list))

#             except TimeoutException:
#                 st.warning(f"⚠️ No products found for brand **{brand}** in category **{selected_category}**")
#                 # When no products are found, ensure N/A is written to Excel
#                 save_data_to_file(row_index, " ", " ", " ", " ") # Save "N/A" for this row
#                 continue
#             except Exception as e:
#                 st.error(f"An error occurred while scraping product tiles: {e}")
#                 save_data_to_file(row_index, " ", " ", " ", " ")
#                 continue

#             # Initialize lists to store multiple matches if fuzzy matching
#             matched_urls = []
#             matched_discounted_prices = []
#             matched_original_prices = []
#             matched_thc_contents = []

#             best_match, best_score = None, 0.0

#             if len(excel_keyword_tokens_set) <= 3:
#                 match_threshold = 0.6  # 60%
#                 print("Threshold set to 60% due to <= 3 Excel tokens.")
#             else:
#                 match_threshold = 0.75 # 75%
#                 print("Threshold set to 75% due to > 3 Excel tokens.")

#             # Update the loop to unpack new scraped fields
#             for name, url, discounted_price, original_price, thc_content, site_weight_tokens, site_qty_num, site_qty_unit, site_flavors, site_keyword_tokens_list in scraped:
#                 # --- QUANTITY COMPARISON LOGIC ---
#                 quantity_match = True
#                 if excel_qty_num is not None and site_qty_num is not None:
#                     if not (excel_qty_num == site_qty_num and excel_qty_unit == site_qty_unit):
#                         quantity_match = False
#                         print(f"  Quantity mismatch: Excel '{excel_qty_num} {excel_qty_unit}' vs Site '{site_qty_num} {site_qty_unit}' for '{name}'")

#                 # if no‐weight category, enforce exact weight match before comparing
#                 weight_enforced_match = True
#                 if website_cat in no_weight_categories:
#                     # Normalize site_weight_tokens for consistent comparison (e.g., '1g', '500mg')
#                     normalized_site_weight_tokens = [normalize_weight(swt) for swt in site_weight_tokens]

#                     # Now, compare the normalized weight from the Excel 'Weight' column
#                     # (which is 'normalized_weight') against the site's normalized weight tokens.
#                     if normalized_weight not in normalized_site_weight_tokens:
#                         weight_enforced_match = False
#                         # Update print statement to show the actual Excel column weight being used
#                         print(f"  Weight mismatch for no-weight category: Excel '{normalized_weight}' vs Site '{', '.join(site_weight_tokens)}' for '{name}'")

#                 # Flavor Matching Logic
#                 flavor_match = True
#                 if excel_flavors: # If there are flavors in the Excel product name
#                     # Check if ALL Excel flavors are present in the site product's flavors
#                     if not all(f in site_flavor_tokens for f in excel_flavor_tokens):
#                         flavor_match = False
#                         print(f"  Flavor mismatch: Excel '{', '.join(excel_flavors)}' vs Site '{', '.join(site_flavors) if site_flavors else 'N/A'}' for '{name}'")
#                 # If Excel product has no specific flavors, then any flavor on site is acceptable.
#                 # The 'flavor_match' remains True in this case, meaning no flavor mismatch prevents a match.

#                 # Convert to set for efficient comparison
#                 lc_site_keyword_tokens_set = set(site_keyword_tokens_list)
#                 site_tokens_display = [t.title() for t in site_keyword_tokens_list] # For display

#                 print(f"⚖️ **Site weight tokens for “{name}”:** {', '.join(site_weight_tokens)}")
#                 print(f"📦 **Site quantity for “{name}”:** {site_qty_num} {site_qty_unit if site_qty_unit else 'N/A'}")
#                 print(f"🎨 **Site flavors for “{name}”:** {', '.join(site_flavors) if site_flavors else 'N/A'}") # Display extracted site flavors
#                 print(f"👁️ **Site tokens for “{name}” (cleaned):** {', '.join(site_tokens_display)}")
#                 print(f"💰 **Site Price for “{name}”:** Discounted: {discounted_price}, Original: {original_price}")
#                 print(f"🌿 **Site THC for “{name}”:** {thc_content}")
#                 print(f"🌐 **Site URL for “{name}”:** {url}")

#                 # compare on lowercase using the cleaned keyword token sets
#                 common = [w for w in excel_keyword_tokens_set if w in lc_site_keyword_tokens_set]
#                 common_tokens_display = [t.title() for t in common]
#                 print(f"🔗 **Common tokens:** {', '.join(common_tokens_display)}")

#                 # compute score based only on keyword tokens
#                 if not excel_keyword_tokens_set:
#                     score = 0.0
#                 else:
#                     score = len(common) / len(excel_keyword_tokens_set)
#                 print(f"      Score for “{name}”: {score:.0%}")

#                 # If quantity matched (or wasn't applicable for strict match) and keyword score is good
#                 # And now, ensure weight also matched if it's a no_weight_category
#                 # Ensure flavors also matched if present in Excel
#                 if quantity_match and weight_enforced_match and flavor_match and score >= match_threshold:
#                     # Instead of just taking the best, collect all valid matches
#                     # This is for the scenario where multiple products could fuzzy match
#                     matched_urls.append(url)
#                     matched_discounted_prices.append(discounted_price)
#                     matched_original_prices.append(original_price)
#                     matched_thc_contents.append(thc_content)

#                     # Update best_match for display purposes if a higher score is found
#                     if score > best_score:
#                         best_match_name = name
#                         best_score = score

#             if matched_urls: # If any matches were found
#                 st.success(f"✅ Matched “{target_name}” → “{best_match_name}” ({best_score:.0%})")
#                 st.write(f"   **URL(s):** {', '.join(matched_urls)}")
#                 st.write(f"   **Price(s):** Discounted: {', '.join(map(str, matched_discounted_prices))} (Original: {', '.join(map(str, matched_original_prices))})")
#                 st.write(f"   **THC(s):** {', '.join(matched_thc_contents)}")
#                 # Save the collected data for this row
#                 save_data_to_file(row_index, matched_discounted_prices, matched_original_prices, matched_thc_contents, matched_urls)
#             else:
#                 st.warning(f"⚠️ No ≥{int(match_threshold * 100)}% match for “{target_name}” (including quantity, weight, and flavor comparisons).")
#                 # When no match, save "N/A" for the current row
#                 save_data_to_file(row_index, " ", " ", " ", " ")
#             # --- PRODUCT MATCHING END ---

#             # Optionally, add a short delay or confirmation after each selection
#             time.sleep(3)

#         st.write("Scraping completed for category:", selected_category)

#         driver.quit()  # Close the driver after both steps are completed

#         # Add the download button after scraping is complete and driver is quit
#         if excel_buffer is not None:
#             st.download_button(
#                 label="Download Updated Excel File",
#                 data=excel_buffer.getvalue(), # Get the BytesIO content
#                 file_name=f"updated_{uploaded_file.name}",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#             )





















